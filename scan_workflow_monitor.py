"""
Automated Swing Scanner Workflow Monitor.

This module owns the complete automated scanner workflow:

1. Start the Streamlit scanner.
2. Open the application through Playwright.
3. Select the requested market segment.
4. Configure the stock scan limit.
5. Trigger the market scan.
6. Wait for scan completion.
7. Download the complete Backtest Track Record.
8. Save the latest scan results.
9. Save dated scan history.
10. Extract today's signalled stocks.
11. Save the latest signal results.
12. Save dated signal history.
13. Generate and log a workflow summary.
14. Stop the managed Streamlit server.

The implementation is intentionally organised into four sections:

Part 1
    - Imports
    - Constants
    - Configuration and result models
    - Logging
    - Path and directory management
    - Streamlit server lifecycle

Part 2
    - Playwright lifecycle
    - Scanner UI readiness
    - Segment selection
    - Stock-limit configuration
    - Market scan triggering
    - Scan completion detection

Part 3
    - Backtest Track Record download
    - CSV loading and validation
    - Signal extraction
    - Atomic result persistence
    - Integrity checks

Part 4
    - Temporary download lifecycle
    - Complete workflow orchestration
    - Workflow summary
    - CLI parsing
    - Application entry point
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import logging
import os
import re
import socket
import subprocess
import sys
import time
import io

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Final
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd

from playwright.sync_api import (
    Browser,
    Error as PlaywrightError,
    Locator,
    Page,
    Playwright,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)


# ============================================================
# LOGGING
# ============================================================

LOGGER_NAME: Final[str] = "scan_workflow_monitor"

logger = logging.getLogger(
    LOGGER_NAME
)


# ============================================================
# PROJECT PATHS
# ============================================================

PROJECT_ROOT: Final[Path] = (
    Path(__file__).resolve().parent
)

APP_PATH: Final[Path] = (
    PROJECT_ROOT
    / "swing_scanner_app.py"
)

DOWNLOAD_ROOT: Final[Path] = (
    PROJECT_ROOT
    / "downloads"
)

LATEST_RESULTS_DIR: Final[Path] = (
    DOWNLOAD_ROOT
    / "latest"
)

HISTORY_RESULTS_DIR: Final[Path] = (
    DOWNLOAD_ROOT
    / "history"
)

TEMPORARY_DOWNLOAD_DIR: Final[Path] = (
    DOWNLOAD_ROOT
    / "_temporary"
)

DEBUG_DIR: Final[Path] = (
    PROJECT_ROOT
    / "workflow_debug"
)


# ============================================================
# FILE AND RESULT CONSTANTS
# ============================================================

RESULT_FILE_EXTENSION: Final[str] = ".csv"

STOCK_COLUMN_NAME: Final[str] = "Stock"

SIGNAL_COLUMN_NAME: Final[str] = "Signals today"

WORKFLOW_DATE_FORMAT: Final[str] = "%Y%m%d"

WORKFLOW_TIMESTAMP_FORMAT: Final[str] = (
    "%Y%m%d_%H%M%S"
)


# ============================================================
# TIMEOUT CONSTANTS
# ============================================================

PAGE_TIMEOUT_MS: Final[int] = 30_000

PAGE_LOAD_TIMEOUT_MS: Final[int] = 60_000

DOWNLOAD_TIMEOUT_MS: Final[int] = 60_000

DOWNLOAD_BUTTON_TIMEOUT_MS: Final[int] = 30_000

SERVER_START_TIMEOUT_SECONDS: Final[int] = 60

UI_READY_TIMEOUT_SECONDS: Final[int] = 60

SERVER_STOP_TIMEOUT_SECONDS: Final[int] = 10

CONTROL_VERIFICATION_TIMEOUT_SECONDS: Final[int] = 30

FILE_READY_TIMEOUT_SECONDS: Final[int] = 30

POLL_INTERVAL_SECONDS: Final[float] = 1.0

CONTROL_POLL_INTERVAL_SECONDS: Final[float] = 0.5


# ============================================================
# BROWSER CONSTANTS
# ============================================================

DEFAULT_BROWSER_VIEWPORT_WIDTH: Final[int] = 1600

DEFAULT_BROWSER_VIEWPORT_HEIGHT: Final[int] = 1200

STREAMLIT_SERVER_HEADLESS: Final[bool] = True


# ============================================================
# CONFIGURATION MODELS
# ============================================================

@dataclass(
    frozen=True,
    slots=True,
)
class WorkflowConfig:
    """
    Runtime configuration for one scanner workflow execution.
    """

    segment: str

    host: str = "127.0.0.1"

    port: int = 8501

    headless: bool = True

    scan_timeout_seconds: int = 1800

    keep_server_running: bool = False

    limit_stocks: int = 25

    def __post_init__(
        self,
    ) -> None:
        """
        Validate runtime configuration.
        """

        normalized_segment = (
            str(
                self.segment
            )
            .strip()
        )

        if not normalized_segment:

            raise ValueError(
                "Segment cannot be empty."
            )

        if not (
            1
            <= self.port
            <= 65_535
        ):

            raise ValueError(
                "Port must be between 1 and 65535."
            )

        if self.limit_stocks <= 0:

            raise ValueError(
                "Stock limit must be greater than zero."
            )

        if self.scan_timeout_seconds <= 0:

            raise ValueError(
                "Scan timeout must be greater than zero."
            )

        if not str(
            self.host
        ).strip():

            raise ValueError(
                "Host cannot be empty."
            )



@dataclass(
    frozen=True,
    slots=True,
)
class WorkflowPaths:
    """
    All persistent and temporary paths for one workflow run.

    Persistent results are stored in one Excel workbook.
    Each workflow execution appends timestamped history sheets.
    """

    results_workbook_path: Path

    temporary_download_path: Path



@dataclass(
    frozen=True,
    slots=True,
)
class WorkflowResult:
    """
    Final result returned after successful workflow execution.
    """

    segment: str

    scanned_stock_count: int

    signalled_stock_count: int

    results_workbook_path: Path

    scan_history_sheet: str

    signal_history_sheet: str




@dataclass(
    frozen=True,
    slots=True,
)
class ManagedServer:
    """
    Represents the Streamlit server used by this workflow.

    `process` is None when the workflow connects to an already
    running Streamlit server.
    """

    process: subprocess.Popen[str] | None

    started_by_workflow: bool


# ============================================================
# LOGGING CONFIGURATION
# ============================================================

def configure_logging() -> None:
    """
    Configure application logging.

    Existing logging handlers are preserved when logging has
    already been configured by a parent process or test runner.
    """

    root_logger = logging.getLogger()

    if root_logger.handlers:

        return

    logging.basicConfig(
        level=logging.INFO,
        format=(
            "%(asctime)s | "
            "%(levelname)-8s | "
            "%(name)s | "
            "%(message)s"
        ),
    )


# ============================================================
# NORMALIZATION HELPERS
# ============================================================

def normalize_segment_name(
    segment: str,
) -> str:
    """
    Convert a scanner segment into a safe file-name component.

    Examples
    --------
    LargeCap
        -> largecap

    Mid Cap
        -> mid_cap

    Small-Cap
        -> small_cap
    """

    normalized = (
        str(
            segment
        )
        .strip()
        .casefold()
    )

    if not normalized:

        raise ValueError(
            "Segment name cannot be empty."
        )

    normalized = re.sub(
        r"[^a-z0-9]+",
        "_",
        normalized,
    )

    normalized = normalized.strip(
        "_"
    )

    if not normalized:

        raise ValueError(
            "Segment name could not be normalized "
            f"into a valid file component: {segment!r}"
        )

    return normalized


def get_result_date(
    now: datetime | None = None,
) -> str:
    """
    Return the workflow result date in YYYYMMDD format.
    """

    timestamp = (
        now
        if now is not None
        else datetime.now()
    )

    return timestamp.strftime(
        WORKFLOW_DATE_FORMAT
    )


def get_workflow_timestamp(
    now: datetime | None = None,
) -> str:
    """
    Return the workflow timestamp in YYYYMMDD_HHMMSS format.
    """

    timestamp = (
        now
        if now is not None
        else datetime.now()
    )

    return timestamp.strftime(
        WORKFLOW_TIMESTAMP_FORMAT
    )


# ============================================================
# OUTPUT DIRECTORY MANAGEMENT
# ============================================================

def ensure_output_directories() -> None:
    """
    Create all directories required by the workflow.
    """

    directories = (
        DOWNLOAD_ROOT,
        LATEST_RESULTS_DIR,
        HISTORY_RESULTS_DIR,
        TEMPORARY_DOWNLOAD_DIR,
        DEBUG_DIR,
    )

    for directory in directories:

        directory.mkdir(
            parents=True,
            exist_ok=True,
        )

    logger.info(
        "Workflow output directories are ready."
    )



# ============================================================
# WORKFLOW PATH GENERATION
# ============================================================


def build_workflow_paths(
    *,
    segment: str,
    result_date: str | None = None,
    timestamp: str | None = None,
) -> WorkflowPaths:
    """
    Build every path required for one workflow execution.

    Persistent files
    ----------------
    downloads/history/
        swing_scanner_<segment>_history.xlsx

    Workbook sheets
    ---------------
    Latest_Scan
    Latest_Signals
    Scan_<YYYYMMDD_HHMMSS>
    Signals_<YYYYMMDD_HHMMSS>

    Temporary files
    ---------------
    downloads/_temporary/
        backtest_<segment>_<YYYYMMDD_HHMMSS>.csv

    Notes
    -----
    `result_date` is retained for backward compatibility with
    existing callers. Historical sheet identity uses `timestamp`.
    """

    ensure_output_directories()

    normalized_segment = (
        normalize_segment_name(
            segment
        )
    )

    timestamp_value = (
        timestamp
        if timestamp is not None
        else get_workflow_timestamp()
    )

    return WorkflowPaths(
        results_workbook_path=(
            HISTORY_RESULTS_DIR
            / (
                f"swing_scanner_"
                f"{normalized_segment}_"
                f"history.xlsx"
            )
        ),
        temporary_download_path=(
            TEMPORARY_DOWNLOAD_DIR
            / (
                f"backtest_{normalized_segment}"
                f"_{timestamp_value}.csv"
            )
        ),
    )



# ============================================================
# SERVER CONNECTIVITY
# ============================================================

def is_port_open(
    *,
    host: str,
    port: int,
) -> bool:
    """
    Return True when a TCP server is reachable on the target port.
    """

    try:

        with socket.socket(
            socket.AF_INET,
            socket.SOCK_STREAM,
        ) as connection:

            connection.settimeout(
                POLL_INTERVAL_SECONDS
            )

            return (
                connection.connect_ex(
                    (
                        host,
                        port,
                    )
                )
                == 0
            )

    except OSError:

        return False


def build_scanner_url(
    config: WorkflowConfig,
) -> str:
    """
    Build the local Streamlit scanner URL.
    """

    return (
        f"http://{config.host}:"
        f"{config.port}"
    )


# ============================================================
# STREAMLIT SERVER LIFECYCLE
# ============================================================

def start_streamlit_server(
    config: WorkflowConfig,
) -> subprocess.Popen[str]:
    """
    Start the Streamlit scanner as a managed subprocess.

    Important:
    `config.headless` controls only the Playwright browser.

    The Streamlit server always runs in headless server mode.
    This prevents the `--headed` CLI option from affecting
    Streamlit startup.
    """

    if not APP_PATH.exists():

        raise FileNotFoundError(
            "Could not find Streamlit scanner app: "
            f"{APP_PATH}"
        )

    command = [
        sys.executable,
        "-m",
        "streamlit",
        "run",
        str(
            APP_PATH
        ),
        "--server.address",
        config.host,
        "--server.port",
        str(
            config.port
        ),
        "--server.headless",
        "true",
        "--browser.gatherUsageStats",
        "false",
    ]

    logger.info(
        "Starting managed Streamlit scanner."
    )

    logger.info(
        "Streamlit command: %s",
        " ".join(
            command
        ),
    )

    try:

        process = subprocess.Popen(
            command,
            cwd=str(
                PROJECT_ROOT
            ),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
        )

    except OSError as error:

        raise RuntimeError(
            "Failed to start Streamlit scanner."
        ) from error

    return process


def wait_for_streamlit_server(
    *,
    config: WorkflowConfig,
    process: subprocess.Popen[str] | None = None,
) -> None:
    """
    Wait until the Streamlit server accepts TCP connections.

    If the managed Streamlit process exits before the server
    becomes ready, capture and report its available output.
    """

    logger.info(
        "Waiting for Streamlit server | URL=%s",
        build_scanner_url(
            config
        ),
    )

    deadline = (
        time.monotonic()
        + SERVER_START_TIMEOUT_SECONDS
    )

    while (
        time.monotonic()
        < deadline
    ):

        if is_port_open(
            host=config.host,
            port=config.port,
        ):

            logger.info(
                "Streamlit server is ready."
            )

            return

        if (
            process is not None
            and process.poll() is not None
        ):

            stdout_output = ""

            stderr_output = ""

            try:

                stdout_output, stderr_output = (
                    process.communicate(
                        timeout=5
                    )
                )

            except subprocess.TimeoutExpired:

                process.kill()

                stdout_output, stderr_output = (
                    process.communicate()
                )

            raise RuntimeError(
                "Streamlit process exited before "
                "the server became ready. "
                f"ExitCode={process.returncode}\n"
                f"STDOUT:\n{stdout_output[-4000:]}\n"
                f"STDERR:\n{stderr_output[-4000:]}"
            )

        time.sleep(
            POLL_INTERVAL_SECONDS
        )

    if (
        process is not None
        and process.poll() is not None
    ):

        stdout_output = ""

        stderr_output = ""

        try:

            stdout_output, stderr_output = (
                process.communicate(
                    timeout=5
                )
            )

        except subprocess.TimeoutExpired:

            process.kill()

            stdout_output, stderr_output = (
                process.communicate()
            )

        raise RuntimeError(
            "Streamlit process stopped while waiting "
            "for startup. "
            f"ExitCode={process.returncode}\n"
            f"STDOUT:\n{stdout_output[-4000:]}\n"
            f"STDERR:\n{stderr_output[-4000:]}"
        )

    raise TimeoutError(
        "Timed out waiting for Streamlit server at "
        f"{config.host}:{config.port} after "
        f"{SERVER_START_TIMEOUT_SECONDS} seconds."
    )



def acquire_streamlit_server(
    config: WorkflowConfig,
) -> ManagedServer:
    """
    Acquire a Streamlit server for the workflow.

    An already-running server on the configured host and port
    is reused. Otherwise, a new managed server is started.
    """

    if is_port_open(
        host=config.host,
        port=config.port,
    ):

        logger.info(
            "Reusing existing Streamlit server | URL=%s",
            build_scanner_url(
                config
            ),
        )

        return ManagedServer(
            process=None,
            started_by_workflow=False,
        )

    process = (
        start_streamlit_server(
            config
        )
    )

    try:

        wait_for_streamlit_server(
            config=config,
            process=process,
        )

    except Exception:

        stop_streamlit_server(
            process
        )

        raise

    return ManagedServer(
        process=process,
        started_by_workflow=True,
    )


def stop_streamlit_server(
    process: subprocess.Popen[str] | None,
) -> None:
    """
    Stop a managed Streamlit process safely.

    Existing external servers are never stopped because they
    are represented by a None process.
    """

    if process is None:

        return

    if process.poll() is not None:

        logger.info(
            "Managed Streamlit process already stopped | "
            "ExitCode=%s",
            process.returncode,
        )

        return

    logger.info(
        "Stopping managed Streamlit server."
    )

    process.terminate()

    try:

        process.wait(
            timeout=SERVER_STOP_TIMEOUT_SECONDS
        )

        logger.info(
            "Managed Streamlit server stopped gracefully."
        )

    except subprocess.TimeoutExpired:

        logger.warning(
            "Streamlit server did not terminate "
            "within %d seconds. Killing process.",
            SERVER_STOP_TIMEOUT_SECONDS,
        )

        process.kill()

        try:

            process.wait(
                timeout=SERVER_STOP_TIMEOUT_SECONDS
            )

        except subprocess.TimeoutExpired as error:

            raise RuntimeError(
                "Could not terminate managed "
                "Streamlit server."
            ) from error


# ============================================================
# DEBUG ARTIFACTS
# ============================================================

def save_debug_screenshot(
    page: Page,
    name: str,
) -> Path:
    """
    Save a full-page Playwright screenshot for diagnostics.

    Screenshot failures are logged but do not hide the original
    workflow failure.
    """

    ensure_output_directories()

    safe_name = re.sub(
        r"[^a-zA-Z0-9_.-]+",
        "_",
        str(
            name
        ),
    ).strip(
        "_"
    )

    if not safe_name:

        safe_name = (
            f"workflow_debug_"
            f"{get_workflow_timestamp()}"
        )

    path = (
        DEBUG_DIR
        / f"{safe_name}.png"
    )

    try:

        page.screenshot(
            path=str(
                path
            ),
            full_page=True,
        )

        logger.info(
            "Debug screenshot saved | File=%s",
            path,
        )

    except PlaywrightError:

        logger.exception(
            "Failed to save debug screenshot | File=%s",
            path,
        )

    return path


# ============================================================
# PART 2
# PLAYWRIGHT LIFECYCLE AND SCANNER UI AUTOMATION
# ============================================================

# ============================================================
# PLAYWRIGHT BROWSER LIFECYCLE
# ============================================================

def create_browser(
    playwright: Playwright,
    *,
    headless: bool,
) -> Browser:
    """
    Launch the Chromium browser used for scanner automation.
    """

    logger.info(
        "Launching Playwright Chromium browser | "
        "Headless=%s",
        headless,
    )

    return playwright.chromium.launch(
        headless=headless,
    )


def create_scanner_page(
    browser: Browser,
) -> Page:
    """
    Create and configure the scanner browser page.
    """

    context = browser.new_context(
        accept_downloads=True,
        viewport={
            "width": DEFAULT_BROWSER_VIEWPORT_WIDTH,
            "height": DEFAULT_BROWSER_VIEWPORT_HEIGHT,
        },
    )

    page = context.new_page()

    page.set_default_timeout(
        PAGE_TIMEOUT_MS
    )

    page.set_default_navigation_timeout(
        PAGE_LOAD_TIMEOUT_MS
    )

    logger.info(
        "Scanner browser page created | "
        "Viewport=%dx%d",
        DEFAULT_BROWSER_VIEWPORT_WIDTH,
        DEFAULT_BROWSER_VIEWPORT_HEIGHT,
    )

    return page


def open_scanner_page(
    page: Page,
    *,
    config: WorkflowConfig,
) -> None:
    """
    Open the Streamlit scanner and wait until its UI is usable.
    """

    url = build_scanner_url(
        config
    )

    logger.info(
        "Opening scanner | URL=%s",
        url,
    )

    try:

        page.goto(
            url,
            wait_until="domcontentloaded",
            timeout=PAGE_LOAD_TIMEOUT_MS,
        )

    except PlaywrightTimeoutError as error:

        raise RuntimeError(
            "Timed out while opening the Streamlit scanner: "
            f"{url}"
        ) from error

    except PlaywrightError as error:

        raise RuntimeError(
            "Failed to open the Streamlit scanner: "
            f"{url}"
        ) from error

    wait_for_scanner_ui(
        page
    )


# ============================================================
# SCANNER UI READINESS
# ============================================================

def wait_for_scanner_ui(
    page: Page,
) -> None:
    """
    Wait until the Streamlit application is visibly usable.

    The application container is used as the baseline readiness
    signal, with a body visibility check as a fallback.
    """

    logger.info(
        "Waiting for scanner UI readiness."
    )

    deadline = (
        time.monotonic()
        + UI_READY_TIMEOUT_SECONDS
    )

    selectors = (
        '[data-testid="stAppViewContainer"]',
        '[data-testid="stApp"]',
        "body",
    )

    while (
        time.monotonic()
        < deadline
    ):

        for selector in selectors:

            try:

                locator = page.locator(
                    selector
                ).first

                if (
                    locator.count() > 0
                    and locator.is_visible()
                ):

                    logger.info(
                        "Scanner UI is ready | "
                        "Selector=%s",
                        selector,
                    )

                    return

            except PlaywrightError:

                continue

        page.wait_for_timeout(
            int(
                POLL_INTERVAL_SECONDS
                * 1_000
            )
        )

    save_debug_screenshot(
        page,
        "scanner_ui_timeout",
    )

    raise PlaywrightTimeoutError(
        "Timed out waiting for the "
        "Streamlit scanner UI."
    )


# ============================================================
# GENERIC LOCATOR HELPERS
# ============================================================

def wait_for_visible_locator(
    locator: Locator,
    *,
    description: str,
    timeout_ms: int = PAGE_TIMEOUT_MS,
) -> Locator:
    """
    Wait for a locator to become visible and return it.
    """

    try:

        locator.wait_for(
            state="visible",
            timeout=timeout_ms,
        )

    except PlaywrightTimeoutError as error:

        raise RuntimeError(
            f"Timed out waiting for {description}."
        ) from error

    except PlaywrightError as error:

        raise RuntimeError(
            f"Failed while waiting for {description}."
        ) from error

    return locator


def get_visible_locator(
    page: Page,
    selectors: tuple[str, ...],
    *,
    description: str,
) -> Locator:
    """
    Return the first visible locator matching one of the
    supplied selectors.
    """

    for selector in selectors:

        try:

            locator = page.locator(
                selector
            )

            count = locator.count()

            for index in range(
                count
            ):

                candidate = locator.nth(
                    index
                )

                if candidate.is_visible():

                    logger.debug(
                        "%s located | Selector=%s | Index=%d",
                        description,
                        selector,
                        index,
                    )

                    return candidate

        except PlaywrightError:

            continue

    raise RuntimeError(
        f"Could not locate {description}."
    )


# ============================================================
# MARKET SEGMENT SELECTION
# ============================================================


def normalize_segment_value(
    value: str,
) -> str:
    """
    Normalize a market segment value for reliable comparison.

    Examples
    --------
    LargeCap   -> largecap
    Large Cap  -> largecap
    large-cap  -> largecap
    """

    return re.sub(
        r"[^a-z0-9]+",
        "",
        str(
            value
        ).casefold(),
    )


def get_segment_widget(
    page: Page,
) -> Locator:
    """
    Locate the Streamlit widget that contains the market segment
    label and its associated interactive control.

    The function searches for a visible label containing a segment
    keyword and then walks up to the nearest Streamlit widget
    container.
    """

    deadline = (
        time.monotonic()
        + CONTROL_VERIFICATION_TIMEOUT_SECONDS
    )

    label_pattern = re.compile(
        r"^\s*(?:market\s*)?segment\s*$",
        re.IGNORECASE,
    )

    while (
        time.monotonic()
        < deadline
    ):

        labels = page.get_by_text(
            label_pattern
        )

        try:

            label_count = labels.count()

            for index in range(
                label_count
            ):

                label = labels.nth(
                    index
                )

                if not label.is_visible():

                    continue

                widget = label.locator(
                    "xpath=ancestor::*"
                    "[@data-testid="
                    "'stSelectbox' "
                    "or @data-testid="
                    "'stRadio' "
                    "or @data-testid="
                    "'stSegmentedControl'][1]"
                )

                if (
                    widget.count() > 0
                    and widget.first.is_visible()
                ):

                    logger.info(
                        "Market segment widget found | "
                        "TestId=%s",
                        (
                            widget.first.get_attribute(
                                "data-testid"
                            )
                            or "unknown"
                        ),
                    )

                    return widget.first

        except PlaywrightError:

            pass

        page.wait_for_timeout(
            int(
                CONTROL_POLL_INTERVAL_SECONDS
                * 1_000
            )
        )

    save_debug_screenshot(
        page,
        "market_segment_widget_not_found",
    )

    raise RuntimeError(
        "Could not locate the market segment widget."
    )


def get_native_select_control(
    widget: Locator,
) -> Locator | None:
    """
    Return a visible native <select> inside the widget when present.
    """

    try:

        select_control = widget.locator(
            "select"
        )

        if (
            select_control.count() > 0
            and select_control.first.is_visible()
        ):

            return select_control.first

    except PlaywrightError:

        pass

    return None


def get_combobox_control(
    widget: Locator,
) -> Locator | None:
    """
    Return a visible ARIA combobox inside the widget when present.
    """

    try:

        combobox = widget.get_by_role(
            "combobox"
        )

        if (
            combobox.count() > 0
            and combobox.first.is_visible()
        ):

            return combobox.first

    except PlaywrightError:

        pass

    return None


def get_radio_controls(
    widget: Locator,
) -> Locator | None:
    """
    Return visible radio controls inside the segment widget.
    """

    try:

        radios = widget.get_by_role(
            "radio"
        )

        if (
            radios.count() > 0
        ):

            for index in range(
                radios.count()
            ):

                radio = radios.nth(
                    index
                )

                if radio.is_visible():

                    return radios

    except PlaywrightError:

        pass

    return None


def select_segment_from_native_select(
    select_control: Locator,
    *,
    segment: str,
) -> bool:
    """
    Select the segment from a native HTML <select> element.
    """

    requested_value = (
        normalize_segment_value(
            segment
        )
    )

    try:

        options = select_control.locator(
            "option"
        )

        for index in range(
            options.count()
        ):

            option = options.nth(
                index
            )

            option_text = (
                option.inner_text()
                .strip()
            )

            option_value = (
                option.get_attribute(
                    "value"
                )
                or ""
            )

            if (
                normalize_segment_value(
                    option_text
                )
                == requested_value
                or normalize_segment_value(
                    option_value
                )
                == requested_value
            ):

                logger.info(
                    "Selecting segment using native select | "
                    "OptionText=%s | OptionValue=%s",
                    option_text,
                    option_value,
                )

                if option_value:

                    select_control.select_option(
                        value=option_value
                    )

                else:

                    select_control.select_option(
                        label=option_text
                    )

                return True

    except PlaywrightError:

        logger.exception(
            "Native select segment selection failed."
        )

    return False


def select_segment_from_combobox(
    page: Page,
    combobox: Locator,
    *,
    segment: str,
) -> bool:
    """
    Select a segment from a Streamlit / ARIA combobox.
    """

    requested_value = (
        normalize_segment_value(
            segment
        )
    )

    try:

        combobox.click(
            timeout=PAGE_TIMEOUT_MS
        )

        page.wait_for_timeout(
            300
        )

        options = page.get_by_role(
            "option"
        )

        option_count = (
            options.count()
        )

        for index in range(
            option_count
        ):

            option = options.nth(
                index
            )

            if not option.is_visible():

                continue

            option_text = (
                option.inner_text(
                    timeout=2_000
                )
                .strip()
            )

            if (
                normalize_segment_value(
                    option_text
                )
                == requested_value
            ):

                logger.info(
                    "Selecting segment from combobox | "
                    "Option=%s",
                    option_text,
                )

                option.click(
                    timeout=PAGE_TIMEOUT_MS
                )

                return True

    except PlaywrightError:

        logger.exception(
            "Combobox segment selection failed."
        )

    return False


def select_segment_from_radios(
    radios: Locator,
    *,
    segment: str,
) -> bool:
    """
    Select a market segment from radio controls.
    """

    requested_value = (
        normalize_segment_value(
            segment
        )
    )

    try:

        radio_count = radios.count()

        for index in range(
            radio_count
        ):

            radio = radios.nth(
                index
            )

            if not radio.is_visible():

                continue

            accessible_name = (
                radio.get_attribute(
                    "aria-label"
                )
                or ""
            )

            value = (
                radio.get_attribute(
                    "value"
                )
                or ""
            )

            radio_text = (
                f"{accessible_name} {value}"
            )

            if (
                normalize_segment_value(
                    radio_text
                )
                == requested_value
                or requested_value
                in normalize_segment_value(
                    radio_text
                )
            ):

                logger.info(
                    "Selecting segment from radio control | "
                    "Segment=%s",
                    segment,
                )

                radio.check(
                    timeout=PAGE_TIMEOUT_MS
                )

                return True

    except PlaywrightError:

        logger.exception(
            "Radio segment selection failed."
        )

    return False


def verify_native_select_segment(
    select_control: Locator,
    *,
    segment: str,
) -> bool:
    """
    Verify the selected option of a native <select>.
    """

    requested_value = (
        normalize_segment_value(
            segment
        )
    )

    try:

        selected_options = (
            select_control.locator(
                "option:checked"
            )
        )

        if (
            selected_options.count()
            == 0
        ):

            return False

        selected = (
            selected_options.first
        )

        selected_text = (
            selected.inner_text()
            .strip()
        )

        selected_value = (
            selected.get_attribute(
                "value"
            )
            or ""
        )

        return (
            normalize_segment_value(
                selected_text
            )
            == requested_value
            or normalize_segment_value(
                selected_value
            )
            == requested_value
        )

    except PlaywrightError:

        return False


def verify_combobox_segment(
    combobox: Locator,
    *,
    segment: str,
) -> bool:
    """
    Verify the currently displayed combobox value.
    """

    requested_value = (
        normalize_segment_value(
            segment
        )
    )

    try:

        text = (
            combobox.inner_text(
                timeout=2_000
            )
            .strip()
        )

        value = (
            combobox.input_value(
                timeout=2_000
            )
            if combobox.evaluate(
                """
                element => (
                    element instanceof HTMLInputElement
                    || element instanceof HTMLSelectElement
                )
                """
            )
            else ""
        )

        combined = (
            f"{text} {value}"
        )

        return (
            requested_value
            in normalize_segment_value(
                combined
            )
        )

    except PlaywrightError:

        return False


def verify_radio_segment(
    radios: Locator,
    *,
    segment: str,
) -> bool:
    """
    Verify that the requested radio option is checked.
    """

    requested_value = (
        normalize_segment_value(
            segment
        )
    )

    try:

        for index in range(
            radios.count()
        ):

            radio = radios.nth(
                index
            )

            if not radio.is_checked():

                continue

            accessible_name = (
                radio.get_attribute(
                    "aria-label"
                )
                or ""
            )

            value = (
                radio.get_attribute(
                    "value"
                )
                or ""
            )

            combined = (
                f"{accessible_name} {value}"
            )

            if (
                requested_value
                in normalize_segment_value(
                    combined
                )
            ):

                return True

    except PlaywrightError:

        return False

    return False


def select_segment(
    page: Page,
    *,
    segment: str,
) -> None:
    """
    Select and verify the requested market segment.

    Selection order
    ---------------
    1. Native HTML select
    2. ARIA / Streamlit combobox
    3. Radio controls

    The function verifies the actual control state after selection.
    """

    normalized_segment = (
        str(
            segment
        ).strip()
    )

    if not normalized_segment:

        raise ValueError(
            "Segment cannot be empty."
        )

    logger.info(
        "Selecting market segment | Segment=%s",
        normalized_segment,
    )

    widget = get_segment_widget(
        page
    )

    # --------------------------------------------------------
    # NATIVE SELECT
    # --------------------------------------------------------

    native_select = (
        get_native_select_control(
            widget
        )
    )

    if native_select is not None:

        if select_segment_from_native_select(
            native_select,
            segment=normalized_segment,
        ):

            page.wait_for_timeout(
                500
            )

            if verify_native_select_segment(
                native_select,
                segment=normalized_segment,
            ):

                logger.info(
                    "Market segment selected and verified | "
                    "Type=native_select | Segment=%s",
                    normalized_segment,
                )

                return

    # --------------------------------------------------------
    # COMBOBOX
    # --------------------------------------------------------

    combobox = (
        get_combobox_control(
            widget
        )
    )

    if combobox is not None:

        if select_segment_from_combobox(
            page,
            combobox,
            segment=normalized_segment,
        ):

            page.wait_for_timeout(
                750
            )

            if verify_combobox_segment(
                combobox,
                segment=normalized_segment,
            ):

                logger.info(
                    "Market segment selected and verified | "
                    "Type=combobox | Segment=%s",
                    normalized_segment,
                )

                return

    # --------------------------------------------------------
    # RADIO CONTROLS
    # --------------------------------------------------------

    radios = get_radio_controls(
        widget
    )

    if radios is not None:

        if select_segment_from_radios(
            radios,
            segment=normalized_segment,
        ):

            page.wait_for_timeout(
                500
            )

            if verify_radio_segment(
                radios,
                segment=normalized_segment,
            ):

                logger.info(
                    "Market segment selected and verified | "
                    "Type=radio | Segment=%s",
                    normalized_segment,
                )

                return

    # --------------------------------------------------------
    # FAILURE DIAGNOSTICS
    # --------------------------------------------------------

    try:

        widget_html = widget.evaluate(
            """
            element => element.outerHTML
            """
        )

        logger.error(
            "Market segment widget HTML:\n%s",
            widget_html[
                :10_000
            ],
        )

    except PlaywrightError:

        logger.exception(
            "Failed to capture market segment widget HTML."
        )

    save_debug_screenshot(
        page,
        "segment_selection_failed",
    )

    raise RuntimeError(
        "Could not select market segment: "
        f"{normalized_segment}"
    )



def wait_for_segment_selection(
    page: Page,
    *,
    segment: str,
) -> None:
    """
    Verify that the requested segment is selected.

    The verification checks visible segment controls and the
    selected state of matching interactive elements.
    """

    deadline = (
        time.monotonic()
        + CONTROL_VERIFICATION_TIMEOUT_SECONDS
    )

    while (
        time.monotonic()
        < deadline
    ):

        try:

            # ------------------------------------------------
            # Check checked radio buttons.
            # ------------------------------------------------

            checked_radio = (
                page.locator(
                    'input[type="radio"]:checked'
                )
            )

            checked_count = (
                checked_radio.count()
            )

            for index in range(
                checked_count
            ):

                radio = checked_radio.nth(
                    index
                )

                value = (
                    radio.get_attribute(
                        "value"
                    )
                    or ""
                )

                aria_label = (
                    radio.get_attribute(
                        "aria-label"
                    )
                    or ""
                )

                combined = (
                    f"{value} {aria_label}"
                ).casefold()

                if (
                    segment.casefold()
                    in combined
                ):

                    logger.info(
                        "Market segment verified | "
                        "Segment=%s",
                        segment,
                    )

                    return

            # ------------------------------------------------
            # Check selected buttons/options.
            # ------------------------------------------------

            selected_elements = (
                page.locator(
                    '[aria-selected="true"], '
                    '[aria-pressed="true"]'
                )
            )

            selected_count = (
                selected_elements.count()
            )

            for index in range(
                selected_count
            ):

                element = selected_elements.nth(
                    index
                )

                if not element.is_visible():

                    continue

                text = (
                    element.inner_text(
                        timeout=2_000
                    )
                    .strip()
                    .casefold()
                )

                if (
                    segment.casefold()
                    in text
                ):

                    logger.info(
                        "Market segment verified | "
                        "Segment=%s",
                        segment,
                    )

                    return

            # ------------------------------------------------
            # Check visible control text as fallback.
            # ------------------------------------------------

            segment_controls = (
                page.locator(
                    '[data-testid="stSelectbox"], '
                    '[data-testid="stRadio"], '
                    '[data-testid="stSegmentedControl"]'
                )
            )

            control_count = (
                segment_controls.count()
            )

            for index in range(
                control_count
            ):

                control = segment_controls.nth(
                    index
                )

                if not control.is_visible():

                    continue

                text = (
                    control.inner_text(
                        timeout=2_000
                    )
                    .strip()
                    .casefold()
                )

                if (
                    segment.casefold()
                    in text
                ):

                    logger.info(
                        "Market segment verified | "
                        "Segment=%s",
                        segment,
                    )

                    return

        except PlaywrightError:

            pass

        page.wait_for_timeout(
            int(
                CONTROL_POLL_INTERVAL_SECONDS
                * 1_000
            )
        )

    save_debug_screenshot(
        page,
        "segment_verification_timeout",
    )

    raise PlaywrightTimeoutError(
        "Timed out verifying selected market segment: "
        f"{segment}"
    )


# ============================================================
# STOCK LIMIT CONFIGURATION
# ============================================================

def get_stock_limit_slider(
    page: Page,
) -> Locator:
    """
    Locate the stock-limit range slider.
    """

    slider = (
        page.locator(
            'input[type="range"]'
        )
        .first
    )

    return wait_for_visible_locator(
        slider,
        description="stock-limit slider",
    )


def get_slider_bounds(
    slider: Locator,
) -> tuple[int, int]:
    """
    Return validated minimum and maximum values for a slider.
    """

    minimum = slider.get_attribute(
        "min"
    )

    maximum = slider.get_attribute(
        "max"
    )

    if minimum is None or maximum is None:

        raise RuntimeError(
            "Could not determine the stock-limit "
            "slider range."
        )

    try:

        minimum_value = int(
            float(
                minimum
            )
        )

        maximum_value = int(
            float(
                maximum
            )
        )

    except ValueError as error:

        raise RuntimeError(
            "Stock-limit slider contains invalid bounds | "
            f"Min={minimum!r} | Max={maximum!r}"
        ) from error

    if minimum_value > maximum_value:

        raise RuntimeError(
            "Stock-limit slider has invalid bounds | "
            f"Min={minimum_value} | "
            f"Max={maximum_value}"
        )

    return (
        minimum_value,
        maximum_value,
    )


def set_stock_limit(
    page: Page,
    *,
    limit: int,
) -> None:
    """
    Set the maximum number of stocks to scan.

    Keyboard interaction is used because Streamlit sliders can
    be more reliable with real keyboard events than direct DOM
    value mutation.
    """

    if limit <= 0:

        raise ValueError(
            "Stock limit must be greater than zero."
        )

    logger.info(
        "Configuring stock limit | Limit=%d",
        limit,
    )

    slider = get_stock_limit_slider(
        page
    )

    minimum_value, maximum_value = (
        get_slider_bounds(
            slider
        )
    )

    if not (
        minimum_value
        <= limit
        <= maximum_value
    ):

        raise ValueError(
            "Requested stock limit is outside the "
            "supported slider range | "
            f"Requested={limit} | "
            f"Min={minimum_value} | "
            f"Max={maximum_value}"
        )

    try:

        current_value = int(
            float(
                slider.input_value()
            )
        )

    except (
        PlaywrightError,
        ValueError,
    ) as error:

        raise RuntimeError(
            "Could not read the current stock limit."
        ) from error

    logger.info(
        "Stock-limit slider state | "
        "Current=%d | Target=%d | "
        "Range=%d-%d",
        current_value,
        limit,
        minimum_value,
        maximum_value,
    )

    if current_value == limit:

        logger.info(
            "Requested stock limit is already configured."
        )

        return

    try:

        slider.scroll_into_view_if_needed(
            timeout=PAGE_TIMEOUT_MS,
        )

        slider.focus(
            timeout=PAGE_TIMEOUT_MS,
        )

        slider.press(
            "Home",
            timeout=PAGE_TIMEOUT_MS,
        )

        steps = (
            limit
            - minimum_value
        )

        for _ in range(
            steps
        ):

            slider.press(
                "ArrowRight",
                timeout=PAGE_TIMEOUT_MS,
            )

    except PlaywrightError as error:

        save_debug_screenshot(
            page,
            "stock_limit_configuration_failed",
        )

        raise RuntimeError(
            "Failed to configure stock limit."
        ) from error

    wait_for_stock_limit(
        page,
        expected_limit=limit,
    )



def set_stock_limit_to_maximum(
    page: Page,
) -> int:
    """
    Set the stock limit to the maximum value currently supported
    by the scanner UI.

    The maximum is read dynamically from the slider after the
    selected market segment has been configured.
    """

    slider = get_stock_limit_slider(
        page
    )

    minimum_value, maximum_value = (
        get_slider_bounds(
            slider
        )
    )

    logger.info(
        "Detected stock-limit range | "
        "Minimum=%d | Maximum=%d",
        minimum_value,
        maximum_value,
    )

    set_stock_limit(
        page,
        limit=maximum_value,
    )

    logger.info(
        "Stock limit configured to maximum | "
        "Maximum=%d",
        maximum_value,
    )

    return maximum_value


def wait_for_stock_limit(
    page: Page,
    *,
    expected_limit: int,
) -> None:
    """
    Verify that the requested stock limit is active.
    """

    deadline = (
        time.monotonic()
        + CONTROL_VERIFICATION_TIMEOUT_SECONDS
    )

    while (
        time.monotonic()
        < deadline
    ):

        try:

            slider = get_stock_limit_slider(
                page
            )

            current_value = int(
                float(
                    slider.input_value()
                )
            )

            if current_value == expected_limit:

                logger.info(
                    "Stock limit verified successfully | "
                    "Limit=%d",
                    expected_limit,
                )

                return

        except (
            PlaywrightError,
            ValueError,
            RuntimeError,
        ):

            pass

        page.wait_for_timeout(
            int(
                CONTROL_POLL_INTERVAL_SECONDS
                * 1_000
            )
        )

    save_debug_screenshot(
        page,
        "stock_limit_verification_timeout",
    )

    raise PlaywrightTimeoutError(
        "Timed out verifying stock limit: "
        f"{expected_limit}"
    )


# ============================================================
# SCAN MARKET BUTTON
# ============================================================

def get_scan_market_button(
    page: Page,
) -> Locator:
    """
    Locate the Scan Market action button.
    """

    button = (
        page.get_by_role(
            "button",
            name=re.compile(
                r"^\s*Scan Market\s*$",
                re.IGNORECASE,
            ),
        )
        .first
    )

    return wait_for_visible_locator(
        button,
        description="Scan Market button",
    )


def start_market_scan(
    page: Page,
) -> None:
    """
    Trigger the market scan.
    """

    logger.info(
        "Starting market scan."
    )

    button = get_scan_market_button(
        page
    )

    try:

        button.scroll_into_view_if_needed(
            timeout=PAGE_TIMEOUT_MS,
        )

        button.click(
            timeout=PAGE_TIMEOUT_MS,
        )

    except PlaywrightError as error:

        save_debug_screenshot(
            page,
            "scan_start_failed",
        )

        raise RuntimeError(
            "Failed to click the Scan Market button."
        ) from error

    logger.info(
        "Market scan triggered successfully."
    )


# ============================================================
# SCAN COMPLETION DETECTION
# ============================================================

def is_scan_completion_available(
    page: Page,
) -> bool:
    """
    Return True when the completed scan result state is available.

    The Backtest Track Record download control is the canonical
    completion signal because it represents the completed
    all-scanned-symbol result set.
    """

    try:

        download_controls = page.locator(
            '[data-testid="stDownloadButton"]'
        )

        count = download_controls.count()

        for index in range(
            count
        ):

            control = download_controls.nth(
                index
            )

            if not control.is_visible():

                continue

            text = (
                control.inner_text(
                    timeout=2_000
                )
                .strip()
                .casefold()
            )

            if (
                "backtest" in text
                and "track" in text
            ):

                return True

    except PlaywrightError:

        return False

    return False


def wait_for_scan_completion(
    page: Page,
    *,
    timeout_seconds: int,
) -> None:
    """
    Wait until the market scan completes.

    Completion is determined by the availability of the
    Backtest Track Record download control.
    """

    if timeout_seconds <= 0:

        raise ValueError(
            "Scan timeout must be greater than zero."
        )

    logger.info(
        "Waiting for market scan completion | "
        "Timeout=%d seconds",
        timeout_seconds,
    )

    started_at = time.monotonic()

    deadline = (
        started_at
        + timeout_seconds
    )

    last_logged_seconds = -30

    while (
        time.monotonic()
        < deadline
    ):

        elapsed_seconds = int(
            time.monotonic()
            - started_at
        )

        if (
            elapsed_seconds
            - last_logged_seconds
            >= 30
        ):

            logger.info(
                "Market scan still running | "
                "Elapsed=%d seconds",
                elapsed_seconds,
            )

            last_logged_seconds = (
                elapsed_seconds
            )

        if is_scan_completion_available(
            page
        ):

            logger.info(
                "Market scan completed successfully | "
                "Elapsed=%d seconds",
                elapsed_seconds,
            )

            return

        page.wait_for_timeout(
            int(
                POLL_INTERVAL_SECONDS
                * 1_000
            )
        )

    save_debug_screenshot(
        page,
        "scan_completion_timeout",
    )

    raise PlaywrightTimeoutError(
        "Timed out waiting for market scan completion "
        f"after {timeout_seconds} seconds."
    )


# ============================================================
# COMPLETE SCANNER UI CONFIGURATION
# ============================================================

def configure_scanner(
    page: Page,
    *,
    config: WorkflowConfig,
) -> None:
    """
    Configure the scanner for the requested workflow execution.

    This keeps the orchestration layer independent of individual
    Streamlit controls.
    """

    select_segment(
        page,
        segment=config.segment,
    )

    maximum_stock_limit = (
        set_stock_limit_to_maximum(
            page
        )
    )

    logger.info(
        "Scanner configuration completed | "
        "Segment=%s | MaximumStockLimit=%d",
        config.segment,
        maximum_stock_limit,
    )



# ============================================================
# COMPLETE MARKET SCAN
# ============================================================

def execute_market_scan(
    page: Page,
    *,
    config: WorkflowConfig,
) -> None:
    """
    Configure and execute the complete market scan.

    This is the high-level UI operation used by the workflow
    orchestrator.
    """

    configure_scanner(
        page,
        config=config,
    )

    start_market_scan(
        page
    )

    wait_for_scan_completion(
        page,
        timeout_seconds=config.scan_timeout_seconds,
    )


# ============================================================
# PART 3
# BACKTEST DOWNLOAD, RESULT PROCESSING AND PERSISTENCE
# ============================================================

# ============================================================
# DOWNLOAD CONTROL DISCOVERY
# ============================================================

def get_backtest_download_control(
    page: Page,
) -> Locator:
    """
    Locate the Backtest Track Record download control.

    Streamlit normally renders download actions inside an
    stDownloadButton container. The visible control text is used
    instead of relying on a fixed DOM position.
    """

    download_containers = page.locator(
        '[data-testid="stDownloadButton"]'
    )

    try:

        count = download_containers.count()

        for index in range(
            count
        ):

            container = download_containers.nth(
                index
            )

            if not container.is_visible():

                continue

            text = (
                container.inner_text(
                    timeout=5_000
                )
                .strip()
            )

            normalized_text = (
                text.casefold()
            )

            if (
                "backtest" in normalized_text
                and "track" in normalized_text
            ):

                button = (
                    container.get_by_role(
                        "button"
                    )
                    .first
                )

                if (
                    button.count() > 0
                    and button.is_visible()
                ):

                    return button

                return container

    except PlaywrightError as error:

        raise RuntimeError(
            "Failed while locating the Backtest Track "
            "Record download control."
        ) from error

    raise RuntimeError(
        "Could not locate the Backtest Track Record "
        "download control."
    )


# ============================================================
# DOWNLOADED FILE VALIDATION
# ============================================================

def wait_for_downloaded_file(
    path: Path,
    *,
    timeout_seconds: int = FILE_READY_TIMEOUT_SECONDS,
) -> Path:
    """
    Wait until a downloaded file exists and has a non-zero size.

    The function also verifies that the file size is stable across
    two consecutive polling cycles before returning it.
    """

    if timeout_seconds <= 0:

        raise ValueError(
            "Download file timeout must be greater than zero."
        )

    deadline = (
        time.monotonic()
        + timeout_seconds
    )

    previous_size: int | None = None

    while (
        time.monotonic()
        < deadline
    ):

        if path.exists() and path.is_file():

            try:

                current_size = path.stat().st_size

            except OSError:

                current_size = 0

            if current_size > 0:

                if (
                    previous_size is not None
                    and current_size == previous_size
                ):

                    logger.info(
                        "Downloaded file is ready | "
                        "File=%s | Bytes=%d",
                        path,
                        current_size,
                    )

                    return path

                previous_size = current_size

        time.sleep(
            CONTROL_POLL_INTERVAL_SECONDS
        )

    raise TimeoutError(
        "Timed out waiting for downloaded file: "
        f"{path}"
    )


# ============================================================
# CANONICAL BACKTEST DOWNLOAD
# ============================================================

def download_backtest_track_record(
    page: Page,
    *,
    destination: Path,
) -> Path:
    """
    Download the complete Backtest Track Record.

    This is the single canonical implementation for the workflow.
    Any existing duplicate download implementations should be
    removed from the original module.
    """

    destination.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    if destination.exists():

        try:

            destination.unlink()

        except OSError as error:

            raise RuntimeError(
                "Could not remove existing temporary "
                f"download file: {destination}"
            ) from error

    logger.info(
        "Downloading Backtest Track Record | "
        "Destination=%s",
        destination,
    )

    control = (
        get_backtest_download_control(
            page
        )
    )

    try:

        with page.expect_download(
            timeout=DOWNLOAD_TIMEOUT_MS,
        ) as download_info:

            control.click(
                timeout=DOWNLOAD_BUTTON_TIMEOUT_MS,
            )

        download = download_info.value

        suggested_filename = (
            download.suggested_filename
        )

        logger.info(
            "Download received | "
            "SuggestedFile=%s",
            suggested_filename,
        )

        download.save_as(
            str(
                destination
            )
        )

    except PlaywrightTimeoutError as error:

        save_debug_screenshot(
            page,
            "backtest_download_timeout",
        )

        raise RuntimeError(
            "Timed out waiting for the Backtest Track "
            "Record download."
        ) from error

    except PlaywrightError as error:

        save_debug_screenshot(
            page,
            "backtest_download_failed",
        )

        raise RuntimeError(
            "Failed to download the Backtest Track Record."
        ) from error

    return wait_for_downloaded_file(
        destination
    )


# ============================================================
# FILE INTEGRITY
# ============================================================

def calculate_file_sha256(
    path: Path,
    *,
    chunk_size: int = 1_048_576,
) -> str:
    """
    Calculate the SHA-256 checksum for a file.
    """

    if chunk_size <= 0:

        raise ValueError(
            "Chunk size must be greater than zero."
        )

    if not path.exists():

        raise FileNotFoundError(
            f"Cannot calculate SHA-256. "
            f"File does not exist: {path}"
        )

    digest = hashlib.sha256()

    try:

        with path.open(
            "rb"
        ) as file_handle:

            while True:

                chunk = file_handle.read(
                    chunk_size
                )

                if not chunk:

                    break

                digest.update(
                    chunk
                )

    except OSError as error:

        raise RuntimeError(
            "Failed to calculate SHA-256 for file: "
            f"{path}"
        ) from error

    return digest.hexdigest()


def log_file_integrity(
    path: Path,
) -> str:
    """
    Calculate and log the SHA-256 checksum for a result file.
    """

    checksum = calculate_file_sha256(
        path
    )

    try:

        size = path.stat().st_size

    except OSError:

        size = -1

    logger.info(
        "File integrity | "
        "File=%s | "
        "Bytes=%d | "
        "SHA256=%s",
        path,
        size,
        checksum,
    )

    return checksum


# ============================================================
# CSV LOADING
# ============================================================

def load_result_dataframe(
    path: Path,
) -> pd.DataFrame:
    """
    Load the downloaded Backtest Track Record into a DataFrame.

    The function rejects empty or structurally invalid files before
    downstream signal extraction begins.
    """

    if not path.exists():

        raise FileNotFoundError(
            "Result file does not exist: "
            f"{path}"
        )

    if not path.is_file():

        raise ValueError(
            "Result path is not a file: "
            f"{path}"
        )

    try:

        dataframe = pd.read_csv(
            path,
        )

    except (
        OSError,
        UnicodeDecodeError,
        pd.errors.ParserError,
    ) as error:

        raise RuntimeError(
            "Failed to read Backtest Track Record CSV: "
            f"{path}"
        ) from error

    if dataframe.empty:

        raise ValueError(
            "Backtest Track Record contains no rows: "
            f"{path}"
        )

    dataframe = (
        dataframe.copy()
    )

    dataframe.columns = [
        str(
            column
        ).strip()
        for column in dataframe.columns
    ]

    logger.info(
        "Result CSV loaded | "
        "Rows=%d | Columns=%d | File=%s",
        len(
            dataframe
        ),
        len(
            dataframe.columns
        ),
        path,
    )

    return dataframe


# ============================================================
# CSV SCHEMA VALIDATION
# ============================================================

def resolve_column_name(
    dataframe: pd.DataFrame,
    *,
    expected_column: str,
) -> str:
    """
    Resolve a DataFrame column using case-insensitive matching.

    This allows the downloaded CSV to differ only in capitalization
    or harmless surrounding whitespace.
    """

    normalized_columns = {
        str(
            column
        )
        .strip()
        .casefold(): str(
            column
        )
        .strip()
        for column in dataframe.columns
    }

    resolved = normalized_columns.get(
        expected_column.casefold()
    )

    if resolved is None:

        available_columns = ", ".join(
            map(
                str,
                dataframe.columns,
            )
        )

        raise ValueError(
            "Required column is missing | "
            f"Expected={expected_column!r} | "
            f"Available=[{available_columns}]"
        )

    return resolved


def validate_result_dataframe(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Validate the downloaded scanner result schema.

    The workflow requires:
    - Stock
    - Signals today

    Additional columns are preserved unchanged.
    """

    if dataframe.empty:

        raise ValueError(
            "Cannot validate an empty result DataFrame."
        )

    stock_column = resolve_column_name(
        dataframe,
        expected_column=STOCK_COLUMN_NAME,
    )

    signal_column = resolve_column_name(
        dataframe,
        expected_column=SIGNAL_COLUMN_NAME,
    )

    validated = (
        dataframe.copy()
    )

    rename_mapping: dict[
        str,
        str,
    ] = {}

    if stock_column != STOCK_COLUMN_NAME:

        rename_mapping[
            stock_column
        ] = STOCK_COLUMN_NAME

    if signal_column != SIGNAL_COLUMN_NAME:

        rename_mapping[
            signal_column
        ] = SIGNAL_COLUMN_NAME

    if rename_mapping:

        validated = validated.rename(
            columns=rename_mapping
        )

    validated = validated.dropna(
        how="all"
    )

    if validated.empty:

        raise ValueError(
            "Result DataFrame contains only empty rows."
        )

    logger.info(
        "Result schema validated | "
        "Rows=%d | Columns=%d",
        len(
            validated
        ),
        len(
            validated.columns
        ),
    )

    return validated


# ============================================================
# SIGNAL NORMALIZATION
# ============================================================

def normalize_signal_value(
    value: object,
) -> str:
    """
    Normalize a signal value for reliable comparison.
    """

    if pd.isna(
        value
    ):

        return ""

    normalized = (
        str(
            value
        )
        .strip()
        .casefold()
    )

    if normalized in {
        "",
        "nan",
        "none",
        "null",
        "n/a",
        "na",
        "-",
    }:

        return ""

    return normalized


def has_active_signal(
    value: object,
) -> bool:
    """
    Determine whether a Signals today value represents an active
    signal.

    Empty values and explicit no-signal representations are
    excluded. Any meaningful remaining value is treated as an
    active scanner signal.
    """

    normalized = (
        normalize_signal_value(
            value
        )
    )

    if not normalized:

        return False

    no_signal_values = {
        "no signal",
        "no signals",
        "no",
        "false",
        "0",
        "neutral",
        "none",
        "not signalled",
        "not signaled",
    }

    return (
        normalized
        not in no_signal_values
    )


# ============================================================
# SIGNAL EXTRACTION
# ============================================================

def extract_signalled_stocks(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Extract rows that contain an active value in Signals today.

    The full original row is preserved so that downstream
    consumers retain all backtest metrics and context.
    """

    validated = dataframe

    signal_mask = (
        validated[
            SIGNAL_COLUMN_NAME
        ]
        .map(
            has_active_signal
        )
        .fillna(
            False
        )
    )

    signalled = (
        validated.loc[
            signal_mask
        ]
        .copy()
    )

    signalled = (
        signalled.sort_values(
            by=STOCK_COLUMN_NAME,
            kind="stable",
            na_position="last",
        )
        .reset_index(
            drop=True
        )
    )

    logger.info(
        "Signal extraction completed | "
        "Scanned=%d | Signalled=%d",
        len(
            validated
        ),
        len(
            signalled
        ),
    )

    return signalled



# ============================================================
# EXCEL WORKBOOK PERSISTENCE
# ============================================================


EXCEL_ENGINE = "openpyxl"

LATEST_SCAN_SHEET_NAME = "Latest_Scan"

LATEST_SIGNAL_SHEET_NAME = "Latest_Signals"

SCAN_HISTORY_SHEET_PREFIX = "Scan"

SIGNAL_HISTORY_SHEET_PREFIX = "Signals"

MAX_EXCEL_SHEET_NAME_LENGTH = 31


def build_history_sheet_name(
    *,
    prefix: str,
    timestamp: str,
) -> str:
    """
    Build a valid timestamped Excel worksheet name.

    Example
    -------
    Scan_20260810_135442
    Signals_20260810_135442
    """

    sheet_name = (
        f"{prefix}_{timestamp}"
    )

    return sheet_name[
        :MAX_EXCEL_SHEET_NAME_LENGTH
    ]


def normalize_dataframe_for_excel(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Return a copy safe for Excel serialization.

    Pandas missing values are converted to None so openpyxl does
    not receive unsupported NA sentinel objects.
    """

    normalized = dataframe.copy()

    normalized = normalized.astype(
        object
    )

    return normalized.where(
        pd.notna(
            normalized
        ),
        None,
    )


def remove_worksheet_if_exists(
    workbook,
    sheet_name: str,
) -> None:
    """
    Remove an existing worksheet.

    Used only for Latest sheets because those sheets represent the
    newest workflow snapshot and are intentionally replaced.
    """

    if sheet_name in workbook.sheetnames:

        worksheet = workbook[
            sheet_name
        ]

        workbook.remove(
            worksheet
        )


def write_dataframe_to_worksheet(
    *,
    workbook,
    dataframe: pd.DataFrame,
    sheet_name: str,
    replace_existing: bool,
) -> None:
    """
    Write a DataFrame into an Excel worksheet.
    """

    safe_sheet_name = sheet_name[
        :MAX_EXCEL_SHEET_NAME_LENGTH
    ]

    if safe_sheet_name in workbook.sheetnames:

        if not replace_existing:

            raise RuntimeError(
                "Refusing to overwrite existing history sheet: "
                f"{safe_sheet_name}"
            )

        remove_worksheet_if_exists(
            workbook,
            safe_sheet_name,
        )

    worksheet = workbook.create_sheet(
        title=safe_sheet_name
    )

    normalized_dataframe = (
        normalize_dataframe_for_excel(
            dataframe
        )
    )

    for row in dataframe_to_rows(
        normalized_dataframe,
        index=False,
        header=True,
    ):

        worksheet.append(
            row
        )

    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1:

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

    for column_cells in worksheet.columns:

        if not column_cells:

            continue

        column_letter = (
            column_cells[0].column_letter
        )

        maximum_length = 0

        for cell in column_cells:

            cell_value = (
                ""
                if cell.value is None
                else str(
                    cell.value
                )
            )

            maximum_length = max(
                maximum_length,
                len(
                    cell_value
                ),
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                maximum_length + 2,
                12,
            ),
            40,
        )


def create_or_load_results_workbook(
    workbook_path: Path,
):
    """
    Create a new Excel workbook or load the existing history
    workbook.
    """

    workbook_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    if workbook_path.exists():

        logger.info(
            "Loading existing Excel history workbook | "
            "File=%s",
            workbook_path,
        )

        return load_workbook(
            workbook_path
        )

    logger.info(
        "Creating Excel history workbook | "
        "File=%s",
        workbook_path,
    )

    from openpyxl import Workbook

    workbook = Workbook()

    default_sheet = workbook.active

    workbook.remove(
        default_sheet
    )

    return workbook


def save_workbook_atomically(
    *,
    workbook,
    destination: Path,
) -> None:
    """
    Save the Excel workbook using atomic replacement.
    """

    destination.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    temporary_path = (
        destination.parent
        / (
            f".{destination.stem}."
            f"{os.getpid()}."
            f"{get_workflow_timestamp()}.tmp.xlsx"
        )
    )

    try:

        workbook.save(
            temporary_path
        )

        temporary_path.replace(
            destination
        )

    except (
        OSError,
        PermissionError,
    ) as error:

        try:

            if temporary_path.exists():

                temporary_path.unlink()

        except OSError:

            logger.exception(
                "Failed to remove temporary Excel workbook | "
                "File=%s",
                temporary_path,
            )

        raise RuntimeError(
            "Failed to persist Excel workbook atomically: "
            f"{destination}"
        ) from error


def validate_results_workbook(
    *,
    workbook_path: Path,
    scan_history_sheet: str,
    signal_history_sheet: str,
    expected_scan_rows: int,
    expected_signal_rows: int,
) -> None:
    """
    Validate workbook existence, required sheets and row counts.
    """

    if not workbook_path.exists():

        raise FileNotFoundError(
            "Excel results workbook was not created: "
            f"{workbook_path}"
        )

    workbook = None

    try:

        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=False,
        )

        required_sheets = {
            LATEST_SCAN_SHEET_NAME,
            LATEST_SIGNAL_SHEET_NAME,
            scan_history_sheet,
            signal_history_sheet,
        }

        missing_sheets = (
            required_sheets
            - set(
                workbook.sheetnames
            )
        )

        if missing_sheets:

            raise RuntimeError(
                "Excel workbook is missing required sheets: "
                f"{sorted(missing_sheets)}"
            )

        scan_sheet = workbook[
            scan_history_sheet
        ]

        signal_sheet = workbook[
            signal_history_sheet
        ]

        actual_scan_rows = max(
            scan_sheet.max_row - 1,
            0,
        )

        actual_signal_rows = max(
            signal_sheet.max_row - 1,
            0,
        )

        if actual_scan_rows != expected_scan_rows:

            raise RuntimeError(
                "Excel scan history row count mismatch | "
                f"Expected={expected_scan_rows} | "
                f"Actual={actual_scan_rows}"
            )

        if actual_signal_rows != expected_signal_rows:

            raise RuntimeError(
                "Excel signal history row count mismatch | "
                f"Expected={expected_signal_rows} | "
                f"Actual={actual_signal_rows}"
            )

    except Exception as error:

        raise RuntimeError(
            "Excel results workbook validation failed: "
            f"{workbook_path}"
        ) from error

    finally:

        if workbook is not None:

            workbook.close()

    logger.info(
        "Excel results workbook validated successfully | "
        "File=%s",
        workbook_path,
    )


# ============================================================
# COMPLETE RESULT PERSISTENCE
# ============================================================

def persist_workflow_results(
    *,
    scan_dataframe: pd.DataFrame,
    signalled_dataframe: pd.DataFrame,
    paths: WorkflowPaths,
) -> tuple[
    str,
    str,
]:
    """
    Persist workflow results into one Excel workbook.

    Latest sheets are replaced:
    - Latest_Scan
    - Latest_Signals

    Historical sheets are appended and preserved:
    - Scan_<timestamp>
    - Signals_<timestamp>
    """

    logger.info(
        "Persisting workflow results to Excel workbook."
    )

    if scan_dataframe is None:

        raise ValueError(
            "scan_dataframe cannot be None."
        )

    if signalled_dataframe is None:

        raise ValueError(
            "signalled_dataframe cannot be None."
        )

    run_timestamp = (
        get_workflow_timestamp()
    )

    scan_history_sheet = (
        build_history_sheet_name(
            prefix=SCAN_HISTORY_SHEET_PREFIX,
            timestamp=run_timestamp,
        )
    )

    signal_history_sheet = (
        build_history_sheet_name(
            prefix=SIGNAL_HISTORY_SHEET_PREFIX,
            timestamp=run_timestamp,
        )
    )

    workbook = (
        create_or_load_results_workbook(
            paths.results_workbook_path
        )
    )

    try:

        # ----------------------------------------------------
        # LATEST COMPLETE SCAN
        # ----------------------------------------------------

        write_dataframe_to_worksheet(
            workbook=workbook,
            dataframe=scan_dataframe,
            sheet_name=LATEST_SCAN_SHEET_NAME,
            replace_existing=True,
        )

        # ----------------------------------------------------
        # LATEST SIGNALLED STOCKS
        # ----------------------------------------------------

        write_dataframe_to_worksheet(
            workbook=workbook,
            dataframe=signalled_dataframe,
            sheet_name=LATEST_SIGNAL_SHEET_NAME,
            replace_existing=True,
        )

        # ----------------------------------------------------
        # HISTORICAL COMPLETE SCAN
        # ----------------------------------------------------

        write_dataframe_to_worksheet(
            workbook=workbook,
            dataframe=scan_dataframe,
            sheet_name=scan_history_sheet,
            replace_existing=False,
        )

        # ----------------------------------------------------
        # HISTORICAL SIGNALLED STOCKS
        # ----------------------------------------------------

        write_dataframe_to_worksheet(
            workbook=workbook,
            dataframe=signalled_dataframe,
            sheet_name=signal_history_sheet,
            replace_existing=False,
        )

        save_workbook_atomically(
            workbook=workbook,
            destination=paths.results_workbook_path,
        )

    finally:

        workbook.close()

    validate_results_workbook(
        workbook_path=paths.results_workbook_path,
        scan_history_sheet=scan_history_sheet,
        signal_history_sheet=signal_history_sheet,
        expected_scan_rows=len(
            scan_dataframe
        ),
        expected_signal_rows=len(
            signalled_dataframe
        ),
    )

    logger.info(
        "Workflow results saved successfully | "
        "Workbook=%s | "
        "ScanSheet=%s | "
        "SignalSheet=%s",
        paths.results_workbook_path,
        scan_history_sheet,
        signal_history_sheet,
    )

    return (
        scan_history_sheet,
        signal_history_sheet,
    )



# ============================================================
# COMPLETE RESULT PROCESSING
# ============================================================

def process_downloaded_results(
    *,
    downloaded_file: Path,
    paths: WorkflowPaths,
) -> tuple[
    pd.DataFrame,
    pd.DataFrame,
    str,
    str,
]:
    """
    Execute the complete result-processing pipeline.

    Pipeline
    --------
    1. Verify downloaded CSV integrity.
    2. Load CSV.
    3. Validate required schema.
    4. Extract active signals.
    5. Persist latest results into Excel sheets.
    6. Append timestamped history sheets.

    Returns
    -------
    tuple[
        pd.DataFrame,
        pd.DataFrame,
        str,
        str,
    ]
        Complete scan results, signalled stocks, scan history sheet
        name and signal history sheet name.
    """

    logger.info(
        "Processing downloaded scanner results | "
        "File=%s",
        downloaded_file,
    )

    log_file_integrity(
        downloaded_file
    )

    raw_dataframe = (
        load_result_dataframe(
            downloaded_file
        )
    )

    scan_dataframe = (
        validate_result_dataframe(
            raw_dataframe
        )
    )

    signalled_dataframe = (
        extract_signalled_stocks(
            scan_dataframe
        )
    )

    (
        scan_history_sheet,
        signal_history_sheet,
    ) = persist_workflow_results(
        scan_dataframe=scan_dataframe,
        signalled_dataframe=signalled_dataframe,
        paths=paths,
    )

    logger.info(
        "Downloaded results processed successfully | "
        "Scanned=%d | "
        "Signalled=%d | "
        "Workbook=%s | "
        "ScanSheet=%s | "
        "SignalSheet=%s",
        len(
            scan_dataframe
        ),
        len(
            signalled_dataframe
        ),
        paths.results_workbook_path,
        scan_history_sheet,
        signal_history_sheet,
    )

    return (
        scan_dataframe,
        signalled_dataframe,
        scan_history_sheet,
        signal_history_sheet,
    )



# ============================================================
# TEMPORARY DOWNLOAD CLEANUP
# ============================================================

def cleanup_temporary_download(
    path: Path,
) -> None:
    """
    Remove the temporary downloaded Backtest Track Record.

    Persistent latest/history result files are never touched.
    """

    try:

        if path.exists():

            path.unlink()

            logger.info(
                "Temporary download removed | "
                "File=%s",
                path,
            )

    except OSError:

        logger.exception(
            "Failed to remove temporary download | "
            "File=%s",
            path,
        )


# ============================================================
# PART 4
# COMPLETE WORKFLOW ORCHESTRATION AND CLI
# ============================================================

# ============================================================
# WORKFLOW EXECUTION
# ============================================================

def run_scan_workflow(
    config: WorkflowConfig,
) -> WorkflowResult:
    """
    Execute the complete automated Swing Scanner workflow.

    Workflow
    --------
    1. Validate configuration.
    2. Build all workflow paths.
    3. Acquire or start the Streamlit server.
    4. Launch Playwright.
    5. Open the scanner.
    6. Configure and execute the market scan.
    7. Download the Backtest Track Record.
    8. Validate and process downloaded results.
    9. Persist latest and historical scan/signal results.
    10. Clean up temporary resources.
    11. Stop the managed Streamlit server unless configured
        to keep it running.

    External Streamlit servers are reused but never stopped.
    """

    configure_logging()

    logger.info(
        "=" * 72
    )

    logger.info(
        "Starting automated Swing Scanner workflow."
    )

    logger.info(
        "Workflow configuration | "
        "Segment=%s | "
        "ScanTimeout=%d | "
        "Host=%s | "
        "Port=%d | "
        "Headless=%s | "
        "KeepServerRunning=%s",
        config.segment,
        config.scan_timeout_seconds,
        config.host,
        config.port,
        config.headless,
        config.keep_server_running,
    )

    paths = build_workflow_paths(
        segment=config.segment
    )

    managed_server: ManagedServer | None = None

    browser: Browser | None = None

    page: Page | None = None

    playwright: Playwright | None = None

    workflow_error: Exception | None = None

    started_at = time.monotonic()

    try:

        # ----------------------------------------------------
        # STEP 1: ACQUIRE STREAMLIT SERVER
        # ----------------------------------------------------

        managed_server = (
            acquire_streamlit_server(
                config
            )
        )

        logger.info(
            "Streamlit server acquired | "
            "StartedByWorkflow=%s",
            managed_server.started_by_workflow,
        )

        # ----------------------------------------------------
        # STEP 2: START PLAYWRIGHT
        # ----------------------------------------------------

        logger.info(
            "Starting Playwright runtime."
        )

        playwright_context = (
            sync_playwright()
        )

        playwright = (
            playwright_context.start()
        )

        browser = create_browser(
            playwright,
            headless=config.headless,
        )

        page = create_scanner_page(
            browser
        )

        # ----------------------------------------------------
        # STEP 3: OPEN SCANNER
        # ----------------------------------------------------

        open_scanner_page(
            page,
            config=config,
        )

        # ----------------------------------------------------
        # STEP 4: CONFIGURE AND EXECUTE SCAN
        # ----------------------------------------------------

        execute_market_scan(
            page,
            config=config,
        )

        # ----------------------------------------------------
        # STEP 5: DOWNLOAD BACKTEST TRACK RECORD
        # ----------------------------------------------------

        downloaded_file = (
            download_backtest_track_record(
                page,
                destination=(
                    paths.temporary_download_path
                ),
            )
        )

        logger.info(
            "Backtest Track Record downloaded | "
            "File=%s",
            downloaded_file,
        )

        # ----------------------------------------------------
        # STEP 6: PROCESS RESULTS
        # ----------------------------------------------------

        (
            scan_dataframe,
            signalled_dataframe,
            scan_history_sheet,
            signal_history_sheet,
        ) = process_downloaded_results(
            downloaded_file=downloaded_file,
            paths=paths,
        )


        # ----------------------------------------------------
        # STEP 7: BUILD FINAL RESULT
        # ----------------------------------------------------

        result = WorkflowResult(
            segment=config.segment,
            scanned_stock_count=len(
                scan_dataframe
            ),
            signalled_stock_count=len(
                signalled_dataframe
            ),
            results_workbook_path=(
                paths.results_workbook_path
            ),
            scan_history_sheet=(
                scan_history_sheet
            ),
            signal_history_sheet=(
                signal_history_sheet
            ),
        )

        elapsed_seconds = (
            time.monotonic()
            - started_at
        )

        log_workflow_success(
            result,
            elapsed_seconds=elapsed_seconds,
        )

        return result

    except Exception as error:

        workflow_error = error

        elapsed_seconds = (
            time.monotonic()
            - started_at
        )

        if page is not None:

            try:

                save_debug_screenshot(
                    page,
                    "workflow_failed",
                )

            except Exception:

                logger.exception(
                    "Failed while creating final "
                    "workflow failure screenshot."
                )

        logger.exception(
            "Swing Scanner workflow failed | "
            "Elapsed=%.2f seconds | "
            "ErrorType=%s | "
            "Error=%s",
            elapsed_seconds,
            type(
                error
            ).__name__,
            error,
        )

        raise

    finally:

        # ----------------------------------------------------
        # CLEANUP TEMPORARY DOWNLOAD
        # ----------------------------------------------------

        cleanup_temporary_download(
            paths.temporary_download_path
        )

        # ----------------------------------------------------
        # CLOSE PLAYWRIGHT BROWSER
        # ----------------------------------------------------

        if browser is not None:

            try:

                logger.info(
                    "Closing Playwright browser."
                )

                browser.close()

            except PlaywrightError:

                logger.exception(
                    "Failed to close Playwright browser."
                )

        # ----------------------------------------------------
        # STOP PLAYWRIGHT RUNTIME
        # ----------------------------------------------------

        if playwright is not None:

            try:

                logger.info(
                    "Stopping Playwright runtime."
                )

                playwright.stop()

            except PlaywrightError:

                logger.exception(
                    "Failed to stop Playwright runtime."
                )

        # ----------------------------------------------------
        # STOP MANAGED STREAMLIT SERVER
        # ----------------------------------------------------

        if (
            managed_server is not None
            and managed_server.started_by_workflow
            and not config.keep_server_running
        ):

            try:

                stop_streamlit_server(
                    managed_server.process
                )

            except Exception:

                logger.exception(
                    "Failed to stop managed "
                    "Streamlit server during cleanup."
                )

        elif (
            managed_server is not None
            and managed_server.started_by_workflow
            and config.keep_server_running
        ):

            logger.info(
                "Keeping managed Streamlit server running "
                "because --keep-server-running was specified."
            )

        elif (
            managed_server is not None
            and not managed_server.started_by_workflow
        ):

            logger.info(
                "External Streamlit server was reused and "
                "will not be stopped."
            )

        # ----------------------------------------------------
        # FINAL WORKFLOW STATE LOG
        # ----------------------------------------------------

        if workflow_error is None:

            logger.info(
                "Workflow cleanup completed successfully."
            )

        else:

            logger.info(
                "Workflow cleanup completed after failure."
            )


# ============================================================
# WORKFLOW SUCCESS SUMMARY
# ============================================================

def log_workflow_success(
    result: WorkflowResult,
    *,
    elapsed_seconds: float,
) -> None:
    """
    Log the final successful workflow summary.
    """

    logger.info(
        "=" * 72
    )

    logger.info(
        "SWING SCANNER WORKFLOW COMPLETED SUCCESSFULLY"
    )

    logger.info(
        "Segment                 : %s",
        result.segment,
    )

    logger.info(
        "Scanned Stocks          : %d",
        result.scanned_stock_count,
    )

    logger.info(
        "Signalled Stocks        : %d",
        result.signalled_stock_count,
    )

    logger.info(
        "Results Workbook        : %s",
        result.results_workbook_path,
    )

    logger.info(
        "Scan History Sheet      : %s",
        result.scan_history_sheet,
    )

    logger.info(
        "Signal History Sheet    : %s",
        result.signal_history_sheet,
    )

    logger.info(
        "Elapsed Time            : %.2f seconds",
        elapsed_seconds,
    )

    logger.info(
        "=" * 72
    )


# ============================================================
# CLI ARGUMENT PARSING
# ============================================================

def parse_arguments(
    argv: list[str] | None = None,
) -> argparse.Namespace:
    """
    Parse command-line arguments.

    Supported arguments
    -------------------
    --segment
        Required scanner market segment.

    --limit
        Maximum number of stocks to scan.

    --timeout
        Maximum time in seconds to wait for scan completion.

    --host
        Streamlit server host.

    --port
        Streamlit server port.

    --headed
        Run Chromium with a visible browser window.

    --keep-server-running
        Do not stop a Streamlit server started by this workflow.
    """

    parser = argparse.ArgumentParser(
        prog="scan_workflow_monitor.py",
        description=(
            "Run the Swing Scanner workflow and save "
            "complete scan and signalled-stock results."
        ),
        formatter_class=(
            argparse.ArgumentDefaultsHelpFormatter
        ),
    )

    parser.add_argument(
        "--segment",
        required=True,
        type=str,
        help=(
            "Market segment to scan."
        ),
    )

    parser.add_argument(
        "--limit",
        dest="limit_stocks",
        type=int,
        default=25,
        help=(
            "Maximum number of stocks to scan."
        ),
    )

    parser.add_argument(
        "--timeout",
        dest="scan_timeout_seconds",
        type=int,
        default=600,
        help=(
            "Maximum seconds to wait for scan completion."
        ),
    )

    parser.add_argument(
        "--host",
        type=str,
        default="127.0.0.1",
        help=(
            "Streamlit server host."
        ),
    )

    parser.add_argument(
        "--port",
        type=int,
        default=8501,
        help=(
            "Streamlit server port."
        ),
    )

    parser.add_argument(
        "--headed",
        action="store_true",
        help=(
            "Run Playwright with a visible browser window."
        ),
    )

    parser.add_argument(
        "--keep-server-running",
        action="store_true",
        help=(
            "Keep a Streamlit server started by this workflow "
            "running after workflow completion."
        ),
    )

    return parser.parse_args(
        argv
    )


# ============================================================
# CLI CONFIGURATION
# ============================================================

def build_workflow_config_from_arguments(
    arguments: argparse.Namespace,
) -> WorkflowConfig:
    """
    Convert parsed CLI arguments into a validated WorkflowConfig.
    """

    return WorkflowConfig(
        segment=arguments.segment,
        limit_stocks=arguments.limit_stocks,
        scan_timeout_seconds=(
            arguments.scan_timeout_seconds
        ),
        host=arguments.host,
        port=arguments.port,
        headless=not arguments.headed,
        keep_server_running=(
            arguments.keep_server_running
        ),
    )



# ============================================================
# RESULT PRINTING
# ============================================================

def print_workflow_result(
    result: WorkflowResult,
) -> None:
    """
    Print a concise machine-readable-friendly workflow summary.

    Logging remains the primary diagnostic output. This summary
    makes CLI execution results immediately visible to the user.
    """

    print()

    print(
        "=" * 72
    )

    print(
        "SWING SCANNER WORKFLOW COMPLETED SUCCESSFULLY"
    )

    print(
        "=" * 72
    )

    print(
        f"Segment: {result.segment}"
    )

    print(
        f"Scanned stocks: "
        f"{result.scanned_stock_count}"
    )

    print(
        f"Signalled stocks: "
        f"{result.signalled_stock_count}"
    )

    print(
        f"Results workbook: "
        f"{result.results_workbook_path}"
    )

    print(
        f"Scan history sheet: "
        f"{result.scan_history_sheet}"
    )

    print(
        f"Signal history sheet: "
        f"{result.signal_history_sheet}"
    )

    print(
        "=" * 72
    )



# ============================================================
# APPLICATION ENTRY POINT
# ============================================================

def main(
    argv: list[str] | None = None,
) -> int:
    """
    Application entry point.

    Returns
    -------
    int
        0 when the workflow completes successfully.
        1 when configuration or workflow execution fails.
    """

    configure_logging()

    try:

        arguments = parse_arguments(
            argv
        )

        config = (
            build_workflow_config_from_arguments(
                arguments
            )
        )

        result = run_scan_workflow(
            config
        )

        print_workflow_result(
            result
        )

        return 0

    except KeyboardInterrupt:

        logger.warning(
            "Workflow interrupted by user."
        )

        return 130

    except (
        ValueError,
        FileNotFoundError,
        RuntimeError,
        TimeoutError,
        PlaywrightError,
    ) as error:

        logger.error(
            "Workflow execution failed | "
            "ErrorType=%s | Error=%s",
            type(
                error
            ).__name__,
            error,
        )

        return 1

    except Exception:

        logger.exception(
            "Unexpected fatal workflow error."
        )

        return 1


# ============================================================
# SCRIPT ENTRY POINT
# ============================================================

if __name__ == "__main__":

    raise SystemExit(
        main()
    )
