"""
===============================================================================
File: excel_writer.py

Description:
    Creates the Daily Excel Report.

Author:
    Pavan Sai
===============================================================================
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter

from config import OUTPUT_FILE

from modules.constants import (
    SheetNames,
    HEADER_FILL,
    WHITE_FONT,
)

from modules.logger import logger

###############################################################################
# Excel Writer
###############################################################################


class ExcelWriter:

    def __init__(self):

        self.workbook = Workbook()

        self.header_fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_FILL,
        )

        self.header_font = Font(
            bold=True,
            color=WHITE_FONT,
        )

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    ###########################################################################

    def write_dataframe(
        self,
        sheet_name: str,
        df: pd.DataFrame,
    ):

        if sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
        else:
            ws = self.workbook.create_sheet(sheet_name)

        # Header

        for col, column in enumerate(df.columns, start=1):

            cell = ws.cell(
                row=1,
                column=col,
                value=column,
            )

            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(
                horizontal="center"
            )

        # Data

        for row in df.itertuples(index=False):

            ws.append(row)

        # Styling

        for row in ws.iter_rows():

            for cell in row:

                cell.border = self.border

        # Auto Width

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value))
                if cell.value is not None else 0
                for cell in column_cells
            )

            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(length + 3, 60)

    ###########################################################################

    def remove_default_sheet(self):

        if "Sheet" in self.workbook.sheetnames:

            del self.workbook["Sheet"]

    ###########################################################################

    def save(
        self,
        output_file: Path = OUTPUT_FILE,
    ):

        self.remove_default_sheet()

        self.workbook.save(output_file)

        logger.info(
            "Excel report saved: %s",
            output_file,
        )


###############################################################################
# Public API
###############################################################################


def create_excel_report(
    gainers: pd.DataFrame,
    losers: pd.DataFrame,
    output_file: Path = OUTPUT_FILE,
):

    writer = ExcelWriter()

    writer.write_dataframe(
        SheetNames.GAINERS.value,
        gainers,
    )

    writer.write_dataframe(
        SheetNames.LOSERS.value,
        losers,
    )

    writer.save(output_file)