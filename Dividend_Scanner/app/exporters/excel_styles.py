"""
Excel Styles

Centralized OpenPyXL styling for reports.
"""

from __future__ import annotations

from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side


class ExcelStyles:

    HEADER_FILL = PatternFill(

        fill_type="solid",

        fgColor="1F4E78",

    )

    HEADER_FONT = Font(

        bold=True,

        color="FFFFFF",

    )

    BORDER = Border(

        left=Side(style="thin"),

        right=Side(style="thin"),

        top=Side(style="thin"),

        bottom=Side(style="thin"),

    )

    CENTER = Alignment(

        horizontal="center",

        vertical="center",

    )

    POSITIVE_FILL = PatternFill(

        fill_type="solid",

        fgColor="C6EFCE",

    )

    NEGATIVE_FILL = PatternFill(

        fill_type="solid",

        fgColor="FFC7CE",

    )

    WARNING_FILL = PatternFill(

        fill_type="solid",

        fgColor="FFEB9C",

    )

    TITLE_FONT = Font(

        bold=True,

        size=14,

    )