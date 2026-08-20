"""
===============================================================================
File: excel_writer.py

Description:
    Creates Latest and History Excel Reports.

Author:
    Pavan Sai
===============================================================================
"""

from __future__ import annotations

from pathlib import Path

from modules.datetime_utils import timestamp

import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from modules.constants import (
    SheetNames,
    HEADER_FILL,
    WHITE_FONT,
)

from modules.logger import logger

###############################################################################
# Output Paths
###############################################################################

OUTPUT_DIR = Path("output")

LATEST_DIR = OUTPUT_DIR / "latest"

HISTORY_DIR = OUTPUT_DIR / "history"

LATEST_DIR.mkdir(
    parents=True,
    exist_ok=True,
)

HISTORY_DIR.mkdir(
    parents=True,
    exist_ok=True,
)

LATEST_FILE = (
    LATEST_DIR /
    "NEWS_Daily_Report.xlsx"
)

HISTORY_FILE = (
    HISTORY_DIR /
    "NEWS_Daily_Report_History.xlsx"
)

###############################################################################
# Excel Writer
###############################################################################


class ExcelWriter:

    def __init__(self):

        self.workbook = Workbook()

        self.header_fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_FILL,
        )

        self.header_font = Font(
            bold=True,
            color=WHITE_FONT,
        )

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    ###########################################################################

    def write_dataframe(
        self,
        sheet_name: str,
        df: pd.DataFrame,
    ):

        if sheet_name in self.workbook.sheetnames:

            ws = self.workbook[sheet_name]

        else:

            ws = self.workbook.create_sheet(
                sheet_name,
            )

        #
        # Header
        #

        for col, column in enumerate(
            df.columns,
            start=1,
        ):

            cell = ws.cell(
                row=1,
                column=col,
                value=column,
            )

            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

            cell.alignment = Alignment(
                horizontal="center",
            )

        #
        # Data
        #

        for row in df.itertuples(
            index=False,
        ):

            ws.append(row)

        #
        # Borders
        #

        for row in ws.iter_rows():

            for cell in row:

                cell.border = self.border

        #
        # Auto Width
        #

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            ws.column_dimensions[
                get_column_letter(
                    column_cells[0].column
                )
            ].width = min(
                length + 3,
                60,
            )

    ###########################################################################

    def remove_default_sheet(self):

        if "Sheet" in self.workbook.sheetnames:

            del self.workbook["Sheet"]

    ###########################################################################

    def save(
        self,
        output_file: Path,
    ):

        self.remove_default_sheet()

        self.workbook.save(
            output_file,
        )

        logger.info(
            "Excel report saved : %s",
            output_file,
        )


###############################################################################
# Latest Report
###############################################################################


def create_latest_report(
    gainers: pd.DataFrame,
    losers: pd.DataFrame,
):

    writer = ExcelWriter()

    writer.write_dataframe(
        SheetNames.GAINERS.value,
        gainers,
    )

    writer.write_dataframe(
        SheetNames.LOSERS.value,
        losers,
    )

    writer.save(
        LATEST_FILE,
    )


###############################################################################
# History Report
###############################################################################


def append_sheet(
    workbook,
    sheet_name: str,
    df: pd.DataFrame,
):

    if sheet_name in workbook.sheetnames:

        ws = workbook[sheet_name]

    else:

        ws = workbook.create_sheet(
            sheet_name,
        )

        ws.append(
            list(df.columns)
        )

    for row in df.itertuples(
        index=False,
    ):

        ws.append(row)


###############################################################################


def update_history_report(
    gainers: pd.DataFrame,
    losers: pd.DataFrame,
):

    if HISTORY_FILE.exists():

        workbook = load_workbook(
            HISTORY_FILE,
        )

    else:

        workbook = Workbook()

        if "Sheet" in workbook.sheetnames:

            workbook.remove(
                workbook["Sheet"]
            )

    append_sheet(
        workbook,
        SheetNames.GAINERS.value,
        gainers,
    )

    append_sheet(
        workbook,
        SheetNames.LOSERS.value,
        losers,
    )

    workbook.save(
        HISTORY_FILE,
    )

    logger.info(
        "History workbook updated : %s",
        HISTORY_FILE,
    )


###############################################################################
# Public API
###############################################################################


def create_excel_report(
    gainers: pd.DataFrame,
    losers: pd.DataFrame,
):

    report_timestamp = timestamp()

    gainers = gainers.copy()

    losers = losers.copy()

    gainers.insert(
        0,
        "Timestamp",
        report_timestamp,
    )

    losers.insert(
        0,
        "Timestamp",
        report_timestamp,
    )

    #
    # Latest Report
    #

    create_latest_report(
        gainers,
        losers,
    )

    #
    # History Report
    #

    update_history_report(
        gainers,
        losers,
    )