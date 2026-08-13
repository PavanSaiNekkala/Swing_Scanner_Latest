from __future__ import annotations

from pathlib import Path

import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook

from config.config import NEWS_UNIVERSE_OUTPUT
from config.logging_config import logger


class NewsExcelExporter:
    """
    Persistent Excel exporter for full-universe news scanning.
    """

    def export(
        self,
        df: pd.DataFrame,
    ) -> Path:

        if not isinstance(
            df,
            pd.DataFrame,
        ):
            raise TypeError(
                "Expected pandas DataFrame."
            )

        if df.empty:
            raise ValueError(
                "News universe dataframe is empty."
            )

        output_path = Path(
            NEWS_UNIVERSE_OUTPUT
        )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        if output_path.exists():

            workbook = load_workbook(
                output_path
            )

        else:

            workbook = Workbook()

            workbook.remove(
                workbook.active
            )

        #######################################################################
        # SHEET
        #######################################################################

        sheet_name = "News"

        if sheet_name in workbook.sheetnames:

            worksheet = workbook[
                sheet_name
            ]

        else:

            worksheet = workbook.create_sheet(
                sheet_name
            )

        #######################################################################
        # HEADERS
        #######################################################################

        if worksheet.max_row == 1 and (
            worksheet["A1"].value is None
        ):

            for column_index, column in enumerate(
                df.columns,
                start=1,
            ):

                worksheet.cell(
                    row=1,
                    column=column_index,
                    value=column,
                )

        #######################################################################
        # APPEND
        #######################################################################

        start_row = (
            worksheet.max_row + 1
        )

        for row_offset, values in enumerate(
            df.itertuples(
                index=False,
                name=None,
            ),
            start=0,
        ):

            excel_row = (
                start_row
                + row_offset
            )

            for column_index, value in enumerate(
                values,
                start=1,
            ):

                if isinstance(
                    value,
                    (list, tuple),
                ):

                    value = " | ".join(
                        map(
                            str,
                            value,
                        )
                    )

                if pd.isna(value):
                    value = None

                worksheet.cell(
                    row=excel_row,
                    column=column_index,
                    value=value,
                )

        worksheet.freeze_panes = "A2"

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        workbook.save(
            output_path
        )

        logger.info(
            "[NEWS EXCEL] Saved %d rows | File=%s",
            len(df),
            output_path,
        )

        return output_path