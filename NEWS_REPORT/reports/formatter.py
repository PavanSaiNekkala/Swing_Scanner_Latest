"""
Institutional Excel Formatter

Responsibilities
----------------
- Centralize workbook styling
- Define reusable NamedStyles
- Apply worksheet formatting
- Format numbers, percentages and currencies
- Configure print settings

This module contains presentation rules only.

Author
------
Pavan Sai Nekkala

Project
-------
NEWS_REPORT

Python
------
3.12+
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook

from config.logging_config import get_logger

logger = get_logger(__name__)

###############################################################################
# Theme
###############################################################################


@dataclass(slots=True, frozen=True)
class Theme:
    """
    Corporate workbook theme.
    """

    HEADER_BG = "1F4E78"

    HEADER_FONT = "FFFFFF"

    BORDER = "D9D9D9"

    POSITIVE = "008000"

    NEGATIVE = "C00000"

    WARNING = "FFB000"

    LINK = "0563C1"

    LIGHT_FILL = "F7F9FB"

###############################################################################
# Formatter
###############################################################################


class ExcelFormatter:
    """
    Institutional formatting engine.
    """

    def __init__(
        self,
        workbook: Workbook,
    ) -> None:

        self.workbook = workbook

        self.styles: dict[
            str,
            NamedStyle,
        ] = {}

        logger.info(
            "ExcelFormatter initialized."
        )

###############################################################################
# Helpers
###############################################################################

    @staticmethod
    def thin_border() -> Border:
        """
        Standard workbook border.
        """

        side = Side(
            style="thin",
            color=Theme.BORDER,
        )

        return Border(
            left=side,
            right=side,
            top=side,
            bottom=side,
        )

###############################################################################
# Header Style
###############################################################################

    def header_style(
        self,
    ) -> NamedStyle:
        """
        Header style.
        """

        style = NamedStyle(
            name="header",
        )

        style.font = Font(
            bold=True,
            size=11,
            color=Theme.HEADER_FONT,
        )

        style.fill = PatternFill(
            fill_type="solid",
            fgColor=Theme.HEADER_BG,
        )

        style.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        style.border = self.thin_border()

        return style

###############################################################################
# Body Style
###############################################################################

    def body_style(
        self,
    ) -> NamedStyle:
        """
        Default body style.
        """

        style = NamedStyle(
            name="body",
        )

        style.font = Font(
            size=10,
        )

        style.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        style.border = self.thin_border()

        return style

###############################################################################
# Percentage Style
###############################################################################

    def percentage_style(
        self,
    ) -> NamedStyle:
        """
        Percentage style.
        """

        style = NamedStyle(
            name="percent",
        )

        style.font = Font(
            size=10,
        )

        style.number_format = "0.00%"

        style.alignment = Alignment(
            horizontal="right",
        )

        style.border = self.thin_border()

        return style

###############################################################################
# Currency Style
###############################################################################

    def currency_style(
        self,
    ) -> NamedStyle:
        """
        Currency style.
        """

        style = NamedStyle(
            name="currency",
        )

        style.font = Font(
            size=10,
        )

        style.number_format = '#,##0.00'

        style.alignment = Alignment(
            horizontal="right",
        )

        style.border = self.thin_border()

        return style

###############################################################################
# Hyperlink Style
###############################################################################

    def hyperlink_style(
        self,
    ) -> NamedStyle:
        """
        Hyperlink style.
        """

        style = NamedStyle(
            name="hyperlink",
        )

        style.font = Font(
            color=Theme.LINK,
            underline="single",
        )

        style.alignment = Alignment(
            vertical="center",
        )

        style.border = self.thin_border()

        return style

###############################################################################
# Register Styles
###############################################################################

    def register_styles(
        self,
    ) -> None:
        """
        Register workbook styles.
        """

        styles = (

            self.header_style(),

            self.body_style(),

            self.percentage_style(),

            self.currency_style(),

            self.hyperlink_style(),

        )

        for style in styles:

            if style.name not in self.workbook.named_styles:

                self.workbook.add_named_style(
                    style,
                )

            self.styles[
                style.name
            ] = style

        logger.info(
            "Workbook styles registered."
        )

###############################################################################
# Worksheet Formatting
###############################################################################

    @staticmethod
    def freeze_panes(
        sheet: Worksheet,
        cell: str = "A2",
    ) -> None:
        """
        Freeze worksheet panes.
        """

        sheet.freeze_panes = cell

    @staticmethod
    def hide_gridlines(
        sheet: Worksheet,
    ) -> None:
        """
        Hide worksheet gridlines.
        """

        sheet.sheet_view.showGridLines = False

    @staticmethod
    def configure_page(
        sheet: Worksheet,
    ) -> None:
        """
        Configure page layout.
        """

        sheet.sheet_view.zoomScale = 90

        sheet.page_setup.orientation = "landscape"

        sheet.page_setup.fitToWidth = 1

        sheet.page_setup.fitToHeight = 0

        sheet.print_options.gridLines = False

        sheet.page_margins.left = 0.30

        sheet.page_margins.right = 0.30

        sheet.page_margins.top = 0.50

        sheet.page_margins.bottom = 0.50

###############################################################################
# Sheet Formatter
###############################################################################

    def format_sheet(
        self,
        sheet: Worksheet,
    ) -> None:
        """
        Apply standard worksheet formatting.
        """

        self.hide_gridlines(
            sheet,
        )

        self.freeze_panes(
            sheet,
        )

        self.configure_page(
            sheet,
        )

###############################################################################
# Workbook Formatter
###############################################################################

    def format_workbook(
        self,
    ) -> None:
        """
        Apply formatting to all worksheets.
        """

        for sheet in self.workbook.worksheets:

            self.format_sheet(
                sheet,
            )

        logger.info(
            "Workbook formatting completed."
        )

###############################################################################
# Auto-fit Columns
###############################################################################

    @staticmethod
    def autofit_columns(
        sheet: Worksheet,
        minimum_width: int = 12,
        maximum_width: int = 60,
    ) -> None:
        """
        Automatically resize worksheet columns.
        """

        from openpyxl.utils import get_column_letter

        for column in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column,
            )

            for cell in column:

                try:

                    if cell.value is None:

                        continue

                    value = str(
                        cell.value,
                    )

                    if len(value) > max_length:

                        max_length = len(
                            value,
                        )

                except Exception:

                    continue

            sheet.column_dimensions[
                column_letter
            ].width = max(
                minimum_width,
                min(
                    max_length + 3,
                    maximum_width,
                ),
            )

###############################################################################
# Hyperlinks
###############################################################################

    @staticmethod
    def apply_hyperlink(
        cell,
        url: str,
    ) -> None:
        """
        Convert a cell into a hyperlink.
        """

        if not url:

            return

        cell.hyperlink = url

        cell.style = "hyperlink"

###############################################################################
# Sentiment Formatting
###############################################################################

    @staticmethod
    def sentiment_fill(
        sentiment: str,
    ) -> PatternFill:
        """
        Background fill for sentiment.
        """

        sentiment = sentiment.lower()

        if sentiment == "bullish":

            return PatternFill(
                fill_type="solid",
                fgColor="C6EFCE",
            )

        if sentiment == "bearish":

            return PatternFill(
                fill_type="solid",
                fgColor="FFC7CE",
            )

        return PatternFill(
            fill_type="solid",
            fgColor="FFF2CC",
        )

###############################################################################
# Apply Sentiment
###############################################################################

    def apply_sentiment(
        self,
        cell,
        sentiment: str,
    ) -> None:
        """
        Format sentiment cell.
        """

        cell.fill = self.sentiment_fill(
            sentiment,
        )

###############################################################################
# Positive / Negative Returns
###############################################################################

    @staticmethod
    def apply_return_font(
        cell,
        value: float,
    ) -> None:
        """
        Apply font color based on return.
        """

        if value > 0:

            cell.font = Font(
                color=Theme.POSITIVE,
                bold=True,
            )

        elif value < 0:

            cell.font = Font(
                color=Theme.NEGATIVE,
                bold=True,
            )

###############################################################################
# Workbook Metadata
###############################################################################

    def configure_metadata(
        self,
    ) -> None:
        """
        Configure workbook metadata.
        """

        props = self.workbook.properties

        props.creator = "Pavan Sai Nekkala"

        props.lastModifiedBy = "NEWS_REPORT"

        props.title = "NSE Market Report"

        props.subject = (
            "Daily NSE Institutional Market Report"
        )

        props.description = (
            "Automatically generated report "
            "containing market movers, "
            "news and AI summaries."
        )

###############################################################################
# Final Formatting
###############################################################################

    def finalize(
        self,
    ) -> None:
        """
        Final formatting for the workbook.
        """

        self.register_styles()

        self.configure_metadata()

        self.format_workbook()

        for sheet in self.workbook.worksheets:

            self.autofit_columns(
                sheet,
            )

        logger.info(
            "Workbook finalized."
        )