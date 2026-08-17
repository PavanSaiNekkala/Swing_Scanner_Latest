"""
Institutional Excel Exporter

Responsibilities
----------------
- Create Excel workbook
- Create worksheets
- Apply professional formatting
- Export MarketReport
- Export NewsReport

This module intentionally DOES NOT

- Fetch market data
- Calculate returns
- Rank news
- Summarize news
- Call AI

Author
------
Pavan Sai Nekkala

Project
-------
NEWS_REPORT

Python
------
3.12+
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config.logging_config import get_logger
from config.settings import settings

from market.models import (
    MarketReport,
    NewsReport,
)

from market.models import ApplicationReport

logger = get_logger(__name__)


###############################################################################
# Excel Exporter
###############################################################################

class ExcelExporter:
    """
    Institutional Excel exporter.
    """

    def __init__(
        self,
    ) -> None:

        self.workbook = Workbook()

        self.styles: dict[
            str,
            NamedStyle,
        ] = {}

        logger.info(
            "ExcelExporter initialized."
        )

###############################################################################
# Workbook
###############################################################################

    def create_workbook(
        self,
    ) -> None:
        """
        Create workbook.
        """

        self.workbook = Workbook()

        default_sheet = self.workbook.active

        self.workbook.remove(
            default_sheet,
        )

        logger.info(
            "Workbook created."
        )

###############################################################################
# Worksheets
###############################################################################

    def create_sheet(
        self,
        title: str,
    ) -> Worksheet:
        """
        Create worksheet.
        """

        sheet = self.workbook.create_sheet(
            title=title,
        )

        sheet.sheet_view.showGridLines = False

        return sheet

###############################################################################
# Styles
###############################################################################

    def create_styles(
        self,
    ) -> None:
        """
        Register reusable styles.
        """

        thin = Side(
            style="thin",
            color="D9D9D9",
        )

        #######################################################################
        # Header
        #######################################################################

        header = NamedStyle(
            name="header",
        )

        header.font = Font(
            bold=True,
            color="FFFFFF",
            size=11,
        )

        header.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        header.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        #######################################################################
        # Body
        #######################################################################

        body = NamedStyle(
            name="body",
        )

        body.font = Font(
            size=10,
        )

        body.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        body.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        #######################################################################
        # Percentage
        #######################################################################

        percent = NamedStyle(
            name="percent",
        )

        percent.font = Font(
            size=10,
        )

        percent.number_format = "0.00%"

        percent.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        #######################################################################
        # Currency
        #######################################################################

        currency = NamedStyle(
            name="currency",
        )

        currency.font = Font(
            size=10,
        )

        currency.number_format = '#,##0.00'

        currency.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        #######################################################################
        # Hyperlink
        #######################################################################

        hyperlink = NamedStyle(
            name="hyperlink",
        )

        hyperlink.font = Font(
            color="0563C1",
            underline="single",
        )

        hyperlink.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        #######################################################################
        # Register Styles
        #######################################################################

        for style in (

            header,

            body,

            percent,

            currency,

            hyperlink,

        ):

            if style.name not in self.workbook.named_styles:

                self.workbook.add_named_style(
                    style,
                )

            self.styles[
                style.name
            ] = style

        logger.info(
            "Workbook styles registered."
        )

###############################################################################
# Helpers
###############################################################################

    def write_cell(
        self,
        sheet: Worksheet,
        row: int,
        column: int,
        value,
        style: str = "body",
    ) -> None:
        """
        Write styled cell.
        """

        cell = sheet.cell(
            row=row,
            column=column,
            value=value,
        )

        cell.style = style

###############################################################################
# Auto Width
###############################################################################

    @staticmethod
    def autofit_columns(
        sheet: Worksheet,
    ) -> None:
        """
        Auto-size worksheet columns.
        """

        for column in sheet.columns:

            width = 0

            letter = get_column_letter(
                column[0].column,
            )

            for cell in column:

                try:

                    width = max(

                        width,

                        len(
                            str(
                                cell.value,
                            )
                        ),

                    )

                except Exception:

                    pass

            sheet.column_dimensions[
                letter
            ].width = min(
                width + 4,
                60,
            )

###############################################################################
# Freeze Panes
###############################################################################

    @staticmethod
    def freeze(
        sheet: Worksheet,
        cell: str = "A2",
    ) -> None:
        """
        Freeze worksheet.
        """

        sheet.freeze_panes = cell

###############################################################################
# Timestamp
###############################################################################

    @staticmethod
    def report_timestamp() -> str:
        """
        Report generation timestamp.
        """

        return datetime.now().strftime(
            "%d-%b-%Y %H:%M:%S",
        )


###############################################################################
# Executive Summary Sheet
###############################################################################

    def write_summary_sheet(
        self,
        market_report: MarketReport,
        news_report: NewsReport,
    ) -> None:
        """
        Create Executive Summary worksheet.
        """

        sheet = self.create_sheet(
            "Executive Summary",
        )

        row = 1

        self.write_cell(
            sheet,
            row,
            1,
            "NSE MARKET REPORT",
            "header",
        )

        row += 2

        summary = [

            (
                "Report Generated",
                self.report_timestamp(),
            ),

            (
                "Universe Size",
                market_report.universe_size,
            ),

            (
                "Top Gainers",
                len(
                    market_report.top_gainers,
                ),
            ),

            (
                "Top Losers",
                len(
                    market_report.top_losers,
                ),
            ),

            (
                "Total News Articles",
                news_report.total_articles,
            ),

        ]

        for metric, value in summary:

            self.write_cell(
                sheet,
                row,
                1,
                metric,
                "header",
            )

            self.write_cell(
                sheet,
                row,
                2,
                value,
            )

            row += 1

        self.autofit_columns(
            sheet,
        )

        self.freeze(
            sheet,
        )

###############################################################################
# Header Writer
###############################################################################

    def write_headers(
        self,
        sheet: Worksheet,
        headers: list[str],
    ) -> None:
        """
        Write worksheet headers.
        """

        for column, header in enumerate(
            headers,
            start=1,
        ):

            self.write_cell(
                sheet,
                1,
                column,
                header,
                "header",
            )

###############################################################################
# Market Sheet
###############################################################################

    def write_market_sheet(
        self,
        title: str,
        stocks,
    ) -> None:
        """
        Write Top Gainers / Top Losers sheet.
        """

        sheet = self.create_sheet(
            title,
        )

        headers = [

            "Rank",

            "Ticker",

            "Company",

            "Sector",

            "Category",

            "Return %",

            "Current Price",

            "Previous Close",

            "Volume",

            "Market Cap",

        ]

        self.write_headers(
            sheet,
            headers,
        )

        row = 2

        for rank, stock in enumerate(
            stocks,
            start=1,
        ):

            self.write_cell(
                sheet,
                row,
                1,
                rank,
            )

            self.write_cell(
                sheet,
                row,
                2,
                stock.identity.ticker,
            )

            self.write_cell(
                sheet,
                row,
                3,
                stock.identity.company_name,
            )

            self.write_cell(
                sheet,
                row,
                4,
                stock.identity.sector,
            )

            self.write_cell(
                sheet,
                row,
                5,
                stock.identity.category,
            )

            self.write_cell(
                sheet,
                row,
                6,
                float(
                    stock.snapshot.day_return_percent
                )
                / 100,
                "percent",
            )

            self.write_cell(
                sheet,
                row,
                7,
                float(
                    stock.snapshot.current_price
                ),
                "currency",
            )

            self.write_cell(
                sheet,
                row,
                8,
                float(
                    stock.snapshot.previous_close
                ),
                "currency",
            )

            self.write_cell(
                sheet,
                row,
                9,
                stock.snapshot.volume,
            )

            self.write_cell(
                sheet,
                row,
                10,
                stock.snapshot.market_cap,
            )

            row += 1

        self.autofit_columns(
            sheet,
        )

        self.freeze(
            sheet,
        )

###############################################################################
# Export Market Sheets
###############################################################################

    def write_market_report(
        self,
        market_report: MarketReport,
    ) -> None:
        """
        Export market worksheets.
        """

        logger.info(
            "Writing market sheets."
        )

        self.write_market_sheet(
            "Top Gainers",
            market_report.top_gainers,
        )

        self.write_market_sheet(
            "Top Losers",
            market_report.top_losers,
        )


###############################################################################
# News Headers
###############################################################################

    def write_news_headers(
        self,
        sheet: Worksheet,
    ) -> None:
        """
        Write headers for News worksheets.
        """

        headers = [

            "Ticker",

            "Company",

            "Sentiment",

            "Confidence",

            "Executive Summary",

            "Headline",

            "Source",

            "Published",

            "Rank Score",

            "URL",

        ]

        self.write_headers(
            sheet,
            headers,
        )

###############################################################################
# One Stock News
###############################################################################

    def write_stock_news(
        self,
        sheet: Worksheet,
        row: int,
        stock,
    ) -> int:
        """
        Write all ranked articles for one stock.

        Returns
        -------
        int
            Next available row.
        """

        first_row = row

        for article in stock.articles:

            ###################################################################
            # Stock Information
            ###################################################################

            self.write_cell(
                sheet,
                row,
                1,
                stock.ticker,
            )

            self.write_cell(
                sheet,
                row,
                2,
                stock.company_name,
            )

            self.write_cell(
                sheet,
                row,
                3,
                stock.sentiment,
            )

            self.write_cell(
                sheet,
                row,
                4,
                stock.confidence,
            )

            ###################################################################
            # Summary only once
            ###################################################################

            if row == first_row:

                self.write_cell(
                    sheet,
                    row,
                    5,
                    stock.executive_summary,
                )

            ###################################################################
            # Article Information
            ###################################################################

            self.write_cell(
                sheet,
                row,
                6,
                article.title,
            )

            self.write_cell(
                sheet,
                row,
                7,
                article.source,
            )

            self.write_cell(
                sheet,
                row,
                8,
                article.published_at.strftime(
                    "%d-%b-%Y %H:%M",
                ),
            )

            self.write_cell(
                sheet,
                row,
                9,
                getattr(
                    article,
                    "rank_score",
                    0.0,
                ),
            )

            ###################################################################
            # Hyperlink
            ###################################################################

            cell = sheet.cell(
                row=row,
                column=10,
                value="Open Article",
            )

            cell.hyperlink = article.url

            cell.style = "hyperlink"

            row += 1

        #######################################################################
        # Blank line between stocks
        #######################################################################

        return row + 1

###############################################################################
# News Worksheet
###############################################################################

    def write_news_sheet(
        self,
        title: str,
        stocks,
    ) -> None:
        """
        Write complete News worksheet.
        """

        sheet = self.create_sheet(
            title,
        )

        self.write_news_headers(
            sheet,
        )

        row = 2

        for stock in stocks:

            row = self.write_stock_news(
                sheet,
                row,
                stock,
            )

        self.autofit_columns(
            sheet,
        )

        self.freeze(
            sheet,
        )

###############################################################################
# Export News
###############################################################################

    def write_news_report(
        self,
        report: NewsReport,
    ) -> None:
        """
        Export all News worksheets.
        """

        logger.info(
            "Writing news worksheets."
        )

        self.write_news_sheet(
            "Gainers News",
            report.gainers_news,
        )

        self.write_news_sheet(
            "Losers News",
            report.losers_news,
        )


###############################################################################
# Workbook Properties
###############################################################################

    def configure_workbook(
        self,
    ) -> None:
        """
        Configure workbook metadata.
        """

        properties = self.workbook.properties

        properties.creator = "Pavan Sai Nekkala"

        properties.lastModifiedBy = "NEWS_REPORT"

        properties.title = "NSE Market Report"

        properties.subject = (
            "Daily NSE Top Gainers & Losers Report"
        )

        properties.description = (
            "Institutional market report containing "
            "top gainers, losers, news, and AI summaries."
        )

###############################################################################
# Worksheet Configuration
###############################################################################

    @staticmethod
    def configure_sheet(
        sheet: Worksheet,
    ) -> None:
        """
        Configure worksheet layout.
        """

        sheet.sheet_view.zoomScale = 90

        sheet.sheet_properties.tabColor = "1F4E78"

        sheet.print_options.gridLines = False

        sheet.page_setup.orientation = "landscape"

        sheet.page_setup.fitToWidth = 1

        sheet.page_setup.fitToHeight = 0

        sheet.page_margins.left = 0.30

        sheet.page_margins.right = 0.30

        sheet.page_margins.top = 0.50

        sheet.page_margins.bottom = 0.50

###############################################################################
# Configure All Worksheets
###############################################################################

    def configure_sheets(
        self,
    ) -> None:
        """
        Configure every worksheet.
        """

        for sheet in self.workbook.worksheets:

            self.configure_sheet(
                sheet,
            )

###############################################################################
# Save Workbook
###############################################################################

    def save(
        self,
        output_file: Path,
    ) -> None:
        """
        Save workbook.
        """

        output_file.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        self.workbook.save(
            output_file,
        )

        logger.info(
            "Workbook saved: %s",
            output_file,
        )

###############################################################################
# Public Export
###############################################################################

    def export(
        self,
        report: ApplicationReport,
        output_file: Path | None = None,
    ) -> Path:
        """
        Export complete Excel report.
        """

        logger.info(
            "=" * 80,
        )

        logger.info(
            "Generating Excel report.",
        )

        logger.info(
            "=" * 80,
        )

        self.create_workbook()

        self.create_styles()

        self.configure_workbook()

        #######################################################################
        # Executive Summary
        #######################################################################

        self.write_summary_sheet(
            report.market,
            report.news,
        )

        #######################################################################
        # Market
        #######################################################################

        self.write_market_report(
            report.market,
        )

        #######################################################################
        # News
        #######################################################################

        self.write_news_report(
            report.news,
        )

        #######################################################################
        # Final Formatting
        #######################################################################

        self.configure_sheets()

        #######################################################################
        # Save
        #######################################################################

        if output_file is None:

            output_file = (
                settings.base.output_directory
                / "NEWS_Report.xlsx"
            )

        self.save(
            output_file,
        )

        logger.info(
            "Excel export completed successfully."
        )

        return output_file


###############################################################################
# Factory
###############################################################################

def build_excel_exporter() -> ExcelExporter:
    """
    Factory for ExcelExporter.
    """

    exporter = ExcelExporter()

    logger.info(
        "ExcelExporter created."
    )

    return exporter