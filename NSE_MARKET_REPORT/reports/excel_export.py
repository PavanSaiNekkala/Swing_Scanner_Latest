"""
reports/excel_export.py

Professional Excel Report Exporter.

Responsibilities
----------------
- Export report DataFrame to Excel.
- Apply professional formatting.
- Create worksheets.
- Apply styles.
- Auto-size columns.
- Freeze panes.
- Apply filters.
- Maintain persistent Gainers / Losers sheets.
- Prevent same-day duplicate appends.
- Save workbook.
"""

from __future__ import annotations

import math

from datetime import datetime

from pathlib import Path

from typing import Any
from typing import Dict

import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side

from openpyxl.utils import get_column_letter

from config.config import (
    DATETIME_FORMAT,
    GAINER_COLOR,
    HEADER_COLOR,
    LOSER_COLOR,
    OUTPUT_EXCEL,
    OUTPUT_EXCEL_DIR,
    REPORT_COLUMNS,
)

from config.logging_config import logger

from config.timezone import now_ist


###############################################################################
# STYLES
###############################################################################

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor=HEADER_COLOR,
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)

CENTER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
)

LEFT_ALIGNMENT = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True,
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


###############################################################################
# EXCEL EXPORTER
###############################################################################

class ExcelExporter:
    """
    Excel export service.
    """

    ###########################################################################
    # INITIALIZATION
    ###########################################################################

    def __init__(
        self,
    ) -> None:

        logger.info(
            "[EXCEL] Initializing exporter..."
        )

        self.timestamp = now_ist()

        OUTPUT_EXCEL_DIR.mkdir(
            parents=True,
            exist_ok=True,
        )

        logger.info(
            "[EXCEL] Ready."
        )

    ###########################################################################
    # DEFAULT FILENAME
    ###########################################################################

    @property
    def filename(
        self,
    ) -> Path:
        """
        Default dated output filename.
        """

        name = self.timestamp.strftime(
            "%Y-%m-%d_NSE_Report.xlsx"
        )

        return (
            OUTPUT_EXCEL_DIR
            / name
        )

    ###########################################################################
    # VALIDATION
    ###########################################################################

    @staticmethod
    def validate_dataframe(
        df: pd.DataFrame,
    ) -> None:
        """
        Validate report dataframe.
        """

        if not isinstance(
            df,
            pd.DataFrame,
        ):

            raise TypeError(
                "Expected pandas DataFrame."
            )

        if df.empty:

            raise ValueError(
                "Report dataframe is empty."
            )

    ###########################################################################
    # COLUMN VALIDATION
    ###########################################################################

    @staticmethod
    def validate_columns(
        df: pd.DataFrame,
    ) -> None:
        """
        Ensure all configured report columns exist.
        """

        missing = [

            column

            for column in REPORT_COLUMNS

            if column not in df.columns

        ]

        if missing:

            raise ValueError(
                "Missing columns: "
                f"{missing}"
            )

    ###########################################################################
    # WORKBOOK
    ###########################################################################

    @staticmethod
    def create_workbook() -> Workbook:
        """
        Create a new workbook.
        """

        return Workbook()

    ###########################################################################
    # WORKSHEET
    ###########################################################################

    @staticmethod
    def create_worksheet(
        workbook: Workbook,
        title: str = "NSE Market Report",
    ):
        """
        Create or rename the active worksheet.
        """

        worksheet = (
            workbook.active
        )

        worksheet.title = title

        return worksheet

    ###########################################################################
    # HEADER
    ###########################################################################

    @staticmethod
    def write_header(
        worksheet,
    ) -> None:
        """
        Write and style report header.
        """

        for column_index, column_name in enumerate(
            REPORT_COLUMNS,
            start=1,
        ):

            cell = worksheet.cell(
                row=1,
                column=column_index,
                value=column_name,
            )

            cell.fill = HEADER_FILL

            cell.font = HEADER_FONT

            cell.alignment = (
                CENTER_ALIGNMENT
            )

            cell.border = THIN_BORDER

    ###########################################################################
    # DATA
    ###########################################################################

    @staticmethod
    def write_data(
        worksheet,
        df: pd.DataFrame,
    ) -> None:
        """
        Write report dataframe rows.
        """

        for row_index, row in enumerate(
            df.itertuples(
                index=False
            ),
            start=2,
        ):

            for column_index, value in enumerate(
                row,
                start=1,
            ):

                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

                cell.border = (
                    THIN_BORDER
                )

                cell.alignment = (
                    LEFT_ALIGNMENT
                )

    ###########################################################################
    # FREEZE
    ###########################################################################

    @staticmethod
    def freeze_header(
        worksheet,
    ) -> None:
        """
        Freeze header row.
        """

        worksheet.freeze_panes = "A2"

    ###########################################################################
    # FILTER
    ###########################################################################

    @staticmethod
    def apply_filter(
        worksheet,
    ) -> None:
        """
        Enable worksheet filter.
        """

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

    ###########################################################################
    # AUTOFIT
    ###########################################################################

    @staticmethod
    def autofit_columns(
        worksheet,
    ) -> None:
        """
        Auto-size worksheet columns.
        """

        for column_cells in (
            worksheet.columns
        ):

            max_length = 0

            column_letter = (
                get_column_letter(
                    column_cells[0].column
                )
            )

            for cell in column_cells:

                try:

                    length = len(
                        str(
                            cell.value
                        )
                    )

                    if (
                        length
                        > max_length
                    ):

                        max_length = length

                except Exception:

                    pass

            adjusted_width = min(
                max_length + 3,
                60,
            )

            worksheet.column_dimensions[
                column_letter
            ].width = (
                adjusted_width
            )

###############################################################################
# NUMBER FORMATS
###############################################################################

    @staticmethod
    def apply_number_formats(
        worksheet,
    ) -> None:
        """
        Apply Excel number formats.
        """

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        price_columns = [

            "CMP",

            "Open",

            "High",

            "Low",

            "Previous Close",

            "Close",

        ]

        for name in price_columns:

            if name not in headers:

                continue

            col = headers[name]

            for row in range(
                2,
                worksheet.max_row + 1,
            ):

                worksheet.cell(
                    row=row,
                    column=col,
                ).number_format = (
                    "#,##0.00"
                )

        #######################################################################
        # CHANGE %
        #######################################################################

        if "1 Day Change %" in headers:

            col = headers[
                "1 Day Change %"
            ]

            for row in range(
                2,
                worksheet.max_row + 1,
            ):

                worksheet.cell(
                    row=row,
                    column=col,
                ).number_format = (
                    "0.00"
                )

        #######################################################################
        # VOLUME
        #######################################################################

        if "Volume" in headers:

            col = headers[
                "Volume"
            ]

            for row in range(
                2,
                worksheet.max_row + 1,
            ):

                worksheet.cell(
                    row=row,
                    column=col,
                ).number_format = (
                    "#,##0"
                )


###############################################################################
# CONDITIONAL FORMATTING
###############################################################################

    @staticmethod
    def apply_conditional_formatting(
        worksheet,
    ) -> None:
        """
        Color rows based on Gainer / Loser category.
        """

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        if "Category" not in headers:

            return

        category_column = headers[
            "Category"
        ]

        gainer_fill = PatternFill(
            fill_type="solid",
            fgColor=GAINER_COLOR,
        )

        loser_fill = PatternFill(
            fill_type="solid",
            fgColor=LOSER_COLOR,
        )

        for row in range(
            2,
            worksheet.max_row + 1,
        ):

            category = (
                worksheet.cell(
                    row=row,
                    column=category_column,
                ).value
            )

            if category is None:

                continue

            category = str(
                category
            ).strip().lower()

            if category == "gainer":

                fill = gainer_fill

            elif category == "loser":

                fill = loser_fill

            else:

                continue

            for col in range(
                1,
                worksheet.max_column + 1,
            ):

                worksheet.cell(
                    row=row,
                    column=col,
                ).fill = fill


###############################################################################
# SUMMARY SHEET
###############################################################################

    def create_summary_sheet(
        self,
        workbook: Workbook,
        df: pd.DataFrame,
    ):
        """
        Create workbook summary worksheet.
        """

        worksheet = (
            workbook.create_sheet(
                title="Summary"
            )
        )

        total = len(df)

        gainers = 0

        losers = 0

        if "Category" in df.columns:

            category_series = (
                df["Category"]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.lower()
            )

            gainers = (
                category_series
                .eq("gainer")
                .sum()
            )

            losers = (
                category_series
                .eq("loser")
                .sum()
            )

        rows = [

            (
                "Report Name",
                "NSE Market Report",
            ),

            (
                "Generated At",
                self.timestamp.strftime(
                    DATETIME_FORMAT
                ),
            ),

            (
                "Total Stocks",
                total,
            ),

            (
                "Top Gainers",
                gainers,
            ),

            (
                "Top Losers",
                losers,
            ),

        ]

        for row_index, (
            label,
            value,
        ) in enumerate(
            rows,
            start=1,
        ):

            label_cell = worksheet.cell(
                row=row_index,
                column=1,
                value=label,
            )

            value_cell = worksheet.cell(
                row=row_index,
                column=2,
                value=value,
            )

            label_cell.font = (
                HEADER_FONT
            )

            label_cell.fill = (
                HEADER_FILL
            )

            label_cell.border = (
                THIN_BORDER
            )

            value_cell.border = (
                THIN_BORDER
            )

        worksheet.column_dimensions[
            "A"
        ].width = 25

        worksheet.column_dimensions[
            "B"
        ].width = 35

        return worksheet


###############################################################################
# DOCUMENT PROPERTIES
###############################################################################

    def apply_document_properties(
        self,
        workbook: Workbook,
    ) -> None:
        """
        Set workbook metadata.
        """

        workbook.properties.creator = (
            "NSE Market Report"
        )

        workbook.properties.title = (
            "NSE Daily Market Report"
        )

        workbook.properties.subject = (
            "Top 20 Gainers & Losers"
        )

        workbook.properties.description = (
            "Automatically generated "
            "NSE Market Report"
        )


###############################################################################
# SAVE WORKBOOK
###############################################################################

    def save(
        self,
        workbook: Workbook,
        output_path: Path | None = None,
    ) -> Path:
        """
        Save workbook to disk.
        """

        if output_path is None:

            output_path = self.filename

        output_path = Path(
            output_path
        )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        if output_path.exists():

            if output_path.is_dir():

                raise IsADirectoryError(
                    "Excel output path is a directory, "
                    f"but a workbook file is required: "
                    f"{output_path}"
                )

            if not output_path.is_file():

                raise OSError(
                    "Excel output path is not "
                    f"a regular file: {output_path}"
                )

        workbook.save(
            output_path
        )

        logger.info(
            "[EXCEL] Saved: %s",
            output_path,
        )

        return output_path

###############################################################################
# APPEND GAINERS / LOSERS
###############################################################################

    def export_gainers_losers_append(
        self,
        df: pd.DataFrame,
    ) -> Path:
        """
        Append the current run to persistent Gainers and Losers sheets.

        Workbook
        --------
        NSE_Market_Report.xlsx

        Sheets
        ------
        Gainers
        Losers

        Duplicate policy
        -----------------
        Existing rows from previous dates are preserved.

        Within the current IST calendar day, an identical row is not
        appended again. Timestamp is deliberately excluded from the
        duplicate signature so multiple executions of the workflow
        during the same day do not create duplicate rows.
        """

        #######################################################################
        # INPUT VALIDATION
        #######################################################################

        if not isinstance(
            df,
            pd.DataFrame,
        ):

            raise TypeError(
                "Expected pandas DataFrame."
            )

        if df.empty:

            raise ValueError(
                "Cannot append an empty market report."
            )

        required_columns = [

            "Timestamp",

            "Category",

            "Symbol",

            "Company",

            "CMP",

            "Open",

            "High",

            "Low",

            "Previous Close",

            "Close",

            "Volume",

            "1 Day Change %",

            "Top Headline",

            "Recent News",

            "Final Remarks",

        ]

        missing = [

            column

            for column in required_columns

            if column not in df.columns

        ]

        if missing:

            raise ValueError(
                "Missing required columns: "
                + ", ".join(missing)
            )

        #######################################################################
        # WORKBOOK PATH
        #######################################################################

        workbook_path = Path(
            OUTPUT_EXCEL
        )

        workbook_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        #######################################################################
        # PATH SAFETY
        #######################################################################

        if workbook_path.exists() and (
            workbook_path.is_dir()
        ):

            raise IsADirectoryError(
                "Excel output path is a directory, "
                f"but a workbook file is required: "
                f"{workbook_path}"
            )

        #######################################################################
        # LOAD / CREATE WORKBOOK
        #######################################################################

        if workbook_path.exists():

            if workbook_path.is_dir():

                raise IsADirectoryError(
                    "Excel output path is a directory, "
                    f"but a workbook file is required: "
                    f"{workbook_path}"
                )

            if not workbook_path.is_file():

                raise OSError(
                    "Excel output path exists but is not "
                    f"a regular file: {workbook_path}"
                )

            workbook = load_workbook(
                workbook_path
            )

        else:

            workbook = Workbook()

            default_sheet = (
                workbook.active
            )

            workbook.remove(
                default_sheet
            )

        #######################################################################
        # SPLIT GAINERS / LOSERS
        #######################################################################

        category_series = (
            df["Category"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )

        gainers = df.loc[
            category_series.eq(
                "gainer"
            )
        ].copy()

        losers = df.loc[
            category_series.eq(
                "loser"
            )
        ].copy()

        #######################################################################
        # NORMALIZATION HELPER
        #######################################################################

        def _normalize_duplicate_value(
            value: Any,
        ) -> str:
            """
            Normalize values for duplicate comparison.
            """

            if value is None:

                return ""

            if isinstance(
                value,
                float,
            ):

                if math.isnan(
                    value
                ):

                    return ""

                return f"{value:.10f}"

            if isinstance(
                value,
                (list, tuple),
            ):

                return " | ".join(
                    str(item).strip()
                    for item in value
                )

            return str(
                value
            ).strip()

        #######################################################################
        # EXISTING SAME-DAY SIGNATURES
        #######################################################################

        def _same_day_duplicate_signatures(
            worksheet,
        ) -> set[tuple]:
            """
            Build duplicate signatures for rows already stored today.
            """

            signatures: set[
                tuple
            ] = set()

            if worksheet.max_row < 2:

                return signatures

            headers = [
                cell.value
                for cell in worksheet[1]
            ]

            header_index = {

                str(header): index

                for index, header
                in enumerate(
                    headers
                )

                if header is not None

            }

            missing_headers = [
                column
                for column in required_columns
                if column not in header_index
            ]

            if missing_headers:

                raise ValueError(
                    "Existing Excel workbook has an "
                    "incompatible schema. Missing headers: "
                    + ", ".join(
                        missing_headers
                    )
                )


            signature_columns = [

                column

                for column in required_columns

                if column != "Timestamp"

            ]

            timestamp_index = (
                header_index.get(
                    "Timestamp"
                )
            )

            if timestamp_index is None:

                return signatures

            today_ist = (
                now_ist().date()
            )

            for row in worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ):

                timestamp_value = row[
                    timestamp_index
                ]

                if timestamp_value is None:

                    continue

                try:

                    if isinstance(
                        timestamp_value,
                        datetime,
                    ):

                        row_date = (
                            timestamp_value.date()
                        )

                    else:

                        timestamp_text = str(
                            timestamp_value
                        ).strip()

                        row_date = (
                            datetime.strptime(
                                timestamp_text[
                                    :10
                                ],
                                "%Y-%m-%d",
                            ).date()
                        )

                except (
                    ValueError,
                    TypeError,
                ):

                    continue

                if row_date != today_ist:

                    continue

                signature = tuple(

                    _normalize_duplicate_value(
                        row[
                            header_index[
                                column
                            ]
                        ]
                    )

                    for column in signature_columns

                    if column
                    in header_index

                )

                signatures.add(
                    signature
                )

            return signatures

        #######################################################################
        # APPEND HELPER
        #######################################################################

        def append_sheet(
            sheet_name: str,
            data: pd.DataFrame,
        ) -> None:
            """
            Append unique rows to one worksheet.
            """

            if sheet_name in (
                workbook.sheetnames
            ):

                worksheet = (
                    workbook[sheet_name]
                )

            else:

                worksheet = (
                    workbook.create_sheet(
                        sheet_name
                    )
                )

            ###################################################################
            # REMOVE EMPTY DEFAULT ROW
            ###################################################################

            if (
                worksheet.max_row == 1
                and all(
                    cell.value is None
                    for cell in worksheet[1]
                )
            ):

                worksheet.delete_rows(
                    1
                )

            ###################################################################
            # HEADER — ALWAYS VALIDATE / REPAIR
            ###################################################################

            expected_headers = list(
                required_columns
            )

            current_headers = [
                worksheet.cell(
                    row=1,
                    column=column_index,
                ).value
                for column_index in range(
                    1,
                    len(expected_headers) + 1,
                )
            ]

            if current_headers != expected_headers:

                for column_index, column in enumerate(
                    expected_headers,
                    start=1,
                ):

                    worksheet.cell(
                        row=1,
                        column=column_index,
                        value=column,
                    )

            ###################################################################
            # HEADER STYLE
            ###################################################################

            for column_index in range(
                1,
                len(expected_headers) + 1,
            ):

                cell = worksheet.cell(
                    row=1,
                    column=column_index,
                )

                cell.font = HEADER_FONT

                cell.fill = HEADER_FILL

                cell.alignment = CENTER_ALIGNMENT

                cell.border = THIN_BORDER

            ###################################################################
            # FREEZE HEADER
            ###################################################################

            worksheet.freeze_panes = "A2"

            ###################################################################
            # SAME-DAY DUPLICATES
            ###################################################################

            existing_signatures = (
                _same_day_duplicate_signatures(
                    worksheet
                )
            )

            signature_columns = [

                column

                for column in required_columns

                if column != "Timestamp"

            ]

            rows_to_append = []

            skipped_duplicates = 0

            for values in data[
                required_columns
            ].itertuples(
                index=False,
                name=None,
            ):

                row_mapping = dict(
                    zip(
                        required_columns,
                        values,
                    )
                )

                signature = tuple(

                    _normalize_duplicate_value(
                        row_mapping.get(
                            column
                        )
                    )

                    for column
                    in signature_columns

                )

                if signature in (
                    existing_signatures
                ):

                    skipped_duplicates += 1

                    continue

                existing_signatures.add(
                    signature
                )

                rows_to_append.append(
                    row_mapping
                )

            ###################################################################
            # NOTHING NEW
            ###################################################################

            if not rows_to_append:

                logger.info(
                    "[EXCEL] %s | "
                    "No new rows to append. "
                    "Skipped duplicates=%d",
                    sheet_name,
                    skipped_duplicates,
                )

                ################################################################
                # FINAL WORKSHEET SETTINGS
                ################################################################

                worksheet.freeze_panes = "A2"

                worksheet.auto_filter.ref = (
                    worksheet.dimensions
                )

                return

            ###################################################################
            # APPEND
            ###################################################################

            start_row = (
                worksheet.max_row + 1
            )

            for row_offset, row_mapping in enumerate(
                rows_to_append,
                start=0,
            ):

                excel_row = (
                    start_row
                    + row_offset
                )

                for column_index, column in enumerate(
                    required_columns,
                    start=1,
                ):

                    value = (
                        row_mapping.get(
                            column
                        )
                    )

                    if isinstance(
                        value,
                        (list, tuple),
                    ):

                        value = (
                            " | ".join(
                                map(
                                    str,
                                    value,
                                )
                            )
                        )

                    if pd.isna(
                        value
                    ):

                        value = None

                    worksheet.cell(
                        row=excel_row,
                        column=column_index,
                        value=value,
                    )

            logger.info(
                "[EXCEL] %s | "
                "Appended=%d | "
                "Skipped duplicates=%d",
                sheet_name,
                len(rows_to_append),
                skipped_duplicates,
            )

            ###################################################################
            # FINAL HEADER STYLE
            ###################################################################

            for column_index in range(
                1,
                len(required_columns) + 1,
            ):

                cell = worksheet.cell(
                    row=1,
                    column=column_index,
                )

                cell.font = HEADER_FONT

                cell.fill = HEADER_FILL

                cell.alignment = CENTER_ALIGNMENT

                cell.border = THIN_BORDER

            ###################################################################
            # FREEZE HEADER
            ###################################################################

            worksheet.freeze_panes = "A2"

            ###################################################################
            # FILTER
            ###################################################################

            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

            ###################################################################
            # WIDTHS
            ###################################################################

            widths = {

                "Timestamp": 22,

                "Category": 12,

                "Symbol": 16,

                "Company": 32,

                "CMP": 14,

                "Open": 14,

                "High": 14,

                "Low": 14,

                "Previous Close": 18,

                "Close": 14,

                "Volume": 16,

                "1 Day Change %": 18,

                "Top Headline": 50,

                "Recent News": 80,

                "Final Remarks": 60,

            }

            for index, column in enumerate(
                required_columns,
                start=1,
            ):

                worksheet.column_dimensions[
                    get_column_letter(
                        index
                    )
                ].width = (
                    widths.get(
                        column,
                        18,
                    )
                )

        #######################################################################
        # APPEND BOTH SHEETS
        #######################################################################

        append_sheet(
            "Gainers",
            gainers,
        )

        append_sheet(
            "Losers",
            losers,
        )

        #######################################################################
        # SAVE
        #######################################################################

        workbook.save(
            workbook_path
        )

        logger.info(
            "[EXCEL] Appended %d gainers and %d losers | File=%s",
            len(gainers),
            len(losers),
            workbook_path,
        )

        return workbook_path

###############################################################################
# STANDARD EXPORT
###############################################################################

    def export(
        self,
        df: pd.DataFrame,
        output_path: Path | None = None,
    ) -> Path:
        """
        Export the formatted report to a standalone Excel workbook.
        """

        logger.info(
            "[EXCEL] Starting Excel export..."
        )

        self.validate_dataframe(
            df
        )

        self.validate_columns(
            df
        )

        workbook = (
            self.create_workbook()
        )

        worksheet = (
            self.create_worksheet(
                workbook
            )
        )

        self.write_header(
            worksheet
        )

        self.write_data(
            worksheet,
            df,
        )

        self.freeze_header(
            worksheet
        )

        self.apply_filter(
            worksheet
        )

        self.apply_number_formats(
            worksheet
        )

        self.autofit_columns(
            worksheet
        )

        self.apply_conditional_formatting(
            worksheet
        )

        self.create_summary_sheet(
            workbook,
            df,
        )

        self.apply_document_properties(
            workbook
        )

        path = self.save(
            workbook,
            output_path,
        )

        logger.info(
            "[EXCEL] Export completed."
        )

        return path


###############################################################################
# EXPORT STATISTICS
###############################################################################

    def statistics(
        self,
        df: pd.DataFrame,
    ) -> Dict[str, Any]:
        """
        Return export statistics.
        """

        self.validate_dataframe(
            df
        )

        return {

            "rows":
                len(df),

            "columns":
                len(df.columns),

            "generated_at":
                self.timestamp.strftime(
                    DATETIME_FORMAT
                ),

            "output_directory":
                str(
                    OUTPUT_EXCEL_DIR
                ),

        }


###############################################################################
# HEALTH CHECK
###############################################################################

    def health_check(
        self,
    ) -> Dict[str, Any]:
        """
        Return exporter health status.
        """

        return {

            "component":
                "ExcelExporter",

            "status":
                "ready",

            "timestamp":
                self.timestamp.strftime(
                    DATETIME_FORMAT
                ),

            "output_directory":
                str(
                    OUTPUT_EXCEL_DIR
                ),

        }


###############################################################################
# RESET
###############################################################################

    def reset(
        self,
    ) -> None:
        """
        Reset exporter timestamp.
        """

        self.timestamp = (
            now_ist()
        )

        logger.info(
            "[EXCEL] Exporter reset."
        )


###############################################################################
# CLOSE
###############################################################################

    def close(
        self,
    ) -> None:
        """
        Cleanup exporter resources.
        """

        logger.info(
            "[EXCEL] Exporter closed."
        )
